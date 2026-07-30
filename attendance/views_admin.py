import csv
import io
import json
import secrets
from datetime import datetime, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Q
from attendance.models import *


def generate_secure_password():
    return secrets.token_urlsafe(8)

@login_required
@transaction.atomic
def manage_streams(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        name = request.POST.get('stream_name', '').strip()
        course_code = request.POST.get('course_code', '').strip()
        
        if name and course_code:
            try:
                course = Course.objects.get(code=course_code)
                Stream.objects.create(name=name, course=course)
                messages.success(request, f"Stream '{name}' successfully registered.")
            except Course.DoesNotExist:
                messages.error(request, "The specified parent course layout does not exist.")
            except Exception as e:
                messages.error(request, f"Error building structural layout: {str(e)}")
        else:
            messages.error(request, "All required input layout items must be filled.")
        return redirect('attendance:manage_streams')

    streams = Stream.objects.select_related('course').all()
    courses = Course.objects.all()
    return render(request, 'attendance/manage_streams.html', {
        'streams': streams,
        'courses': courses
    })

@login_required
@transaction.atomic
def edit_stream(request, stream_id):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    stream_obj = get_object_or_404(Stream, id=stream_id)
    if request.method == 'POST':
        name = request.POST.get('stream_name', '').strip()
        course_code = request.POST.get('course_code', '').strip()
        
        if name and course_code:
            try:
                course = Course.objects.get(code=course_code)
                stream_obj.name = name
                stream_obj.course = course
                stream_obj.save()
                messages.success(request, "Structural stream modifications compiled securely.")
                return redirect('attendance:manage_streams')
            except Course.DoesNotExist:
                messages.error(request, "Target parent course layout context does not exist.")
        else:
            messages.error(request, "Fields cannot be blank.")

    courses = Course.objects.all()
    return render(request, 'attendance/edit_stream.html', {
        'stream': stream_obj,
        'courses': courses
    })

@login_required
@transaction.atomic
def delete_stream(request, stream_id):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    stream_obj = get_object_or_404(Stream, id=stream_id)
    name = stream_obj.name
    stream_obj.delete()
    messages.success(request, f"Stream framework structure '{name}' detached successfully.")
    return redirect('attendance:manage_streams')

@login_required
@transaction.atomic
def bulk_upload_streams(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        if not csv_file.name.endswith('.csv'):
            messages.error(request, "Data error: Expected standard comma-delimited flat CSV file structure.")
            return redirect('attendance:bulk_upload_streams')

        try:
            stream_data = csv_file.read().decode('utf-8')
            io_string = io.StringIO(stream_data)
            reader = csv.reader(io_string)
            next(reader, None)

            success_count = 0
            for row in reader:
                if not row or len(row) < 2:
                    continue
                stream_name = row[0].strip()
                course_code = row[1].strip()

                if stream_name and course_code:
                    course = Course.objects.filter(code=course_code).first()
                    if course:
                        Stream.objects.get_or_create(name=stream_name, course=course)
                        success_count += 1

            messages.success(request, f"Ingested {success_count} unique structural academic stream mappings.")
            return redirect('attendance:manage_streams')
        except Exception as e:
            messages.error(request, f"Parser exception detected: {str(e)}")

    return render(request, 'attendance/bulk_upload_streams.html')

# =========================================================================
# 1. TEACHERS MANAGEMENT
# =========================================================================

@login_required
@transaction.atomic
def manage_teachers(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'single' or ('name' in request.POST and 'email' in request.POST):
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
            course_codes = request.POST.getlist('courses')
            
            if name and email:
                username = email.split('@')[0]
                password = generate_secure_password()
                user, created = User.objects.get_or_create(email=email, defaults={
                    'username': username,
                    'role': User.IS_TEACHER,
                    'raw_password_archive': password
                })
                if created:
                    user.set_password(password)
                    user.save()
                    teacher = TeacherProfile.objects.create(user=user, name=name)
                    if course_codes:
                        teacher.courses.set(Course.objects.filter(code__in=course_codes))
                    messages.success(request, f"Teacher account created for {name}.")
                else:
                    messages.warning(request, "A user with this email already exists.")
            return redirect('attendance:manage_teachers')

        elif action == 'bulk' and request.FILES.get('csv_file'):
            csv_file = request.FILES['csv_file']
            decoded_file = csv_file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(decoded_file), delimiter=',')
            next(reader, None)

            for row in reader:
                if len(row) >= 2:
                    name, email = row[0].strip(), row[1].strip()
                    username = email.split('@')[0]
                    password = generate_secure_password()
                    user, created = User.objects.get_or_create(email=email, defaults={
                        'username': username,
                        'role': User.IS_TEACHER,
                        'raw_password_archive': password
                    })
                    if created:
                        user.set_password(password)
                        user.save()
                        TeacherProfile.objects.create(user=user, name=name)
            return redirect('attendance:export_credentials', role_type='teachers')

    teachers = TeacherProfile.objects.select_related('user').prefetch_related('courses').all()
    all_courses = Course.objects.all()
    return render(request, 'attendance/manage_teachers.html', {
        'teachers': teachers, 
        'all_courses': all_courses,
        'active_courses_list': all_courses
    })


@login_required
@transaction.atomic
def edit_teacher(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    teacher = get_object_or_404(TeacherProfile, pk=pk)
    if request.method == 'POST':
        teacher.name = request.POST.get('name', '').strip()
        teacher.user.email = request.POST.get('email', '').strip()
        teacher.user.save()
        teacher.save()
        
        course_codes = request.POST.getlist('courses')
        teacher.courses.set(Course.objects.filter(code__in=course_codes))
        messages.success(request, "Teacher record updated successfully.")
        return redirect('attendance:manage_teachers')
    
    all_courses = Course.objects.all()
    return render(request, 'attendance/edit_teacher.html', {'teacher': teacher, 'all_courses': all_courses})


@login_required
@transaction.atomic
def delete_teacher(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    teacher = get_object_or_404(TeacherProfile, pk=pk)
    user = teacher.user
    teacher.delete()
    user.delete()
    messages.success(request, "Teacher record permanently deleted.")
    return redirect('attendance:manage_teachers')


# =========================================================================
# 2. STUDENTS MANAGEMENT
# =========================================================================

@login_required
@transaction.atomic
def manage_students(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'single':
            name = (request.POST.get('students_name') or request.POST.get('name', '')).strip()
            reg_number = (request.POST.get('registration_number') or request.POST.get('reg_number', '')).strip()
            email = request.POST.get('email', '').strip()
            course_code = request.POST.get('course', '').strip() 
            stream_id = request.POST.get('stream', '').strip()
            
            if name and reg_number and email and course_code:
                try:
                    course = Course.objects.get(code=course_code)
                    stream = Stream.objects.get(id=stream_id) if stream_id else None
                    password = generate_secure_password()
                    username = email.split('@')[0]
                    
                    user = User.objects.create_user(
                        username=username, 
                        email=email, 
                        password=password,
                        role=User.IS_STUDENT
                    )
                    user.raw_password_archive = password
                    user.save()
                    
                    StudentProfile.objects.create(
                        reg_number=reg_number,
                        user=user,
                        name=name,
                        course=course,
                        stream=stream
                    )
                    messages.success(request, f"Student {name} successfully registered.")
                except Course.DoesNotExist:
                    messages.error(request, "Selected course does not exist.")
                except Stream.DoesNotExist:
                    messages.error(request, "Selected stream does not exist.")
                except Exception as e:
                    messages.error(request, f"Error occurred: {str(e)}")
            
            return redirect('attendance:manage_students')

    students = StudentProfile.objects.select_related('user', 'course', 'stream').all()
    courses = Course.objects.all()
    streams = Stream.objects.all()
    
    return render(request, 'attendance/manage_students.html', {
        'students': students,
        'active_courses_list': courses,
        'active_streams_list': streams
    })

@login_required
@transaction.atomic
def edit_student(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    student = get_object_or_404(StudentProfile, pk=pk)
    if request.method == 'POST':
        student.name = (request.POST.get('students_name') or request.POST.get('name', '')).strip()
        student.reg_number = (request.POST.get('registration_number') or request.POST.get('reg_number', '')).strip()
        student.user.email = request.POST.get('email', '').strip()
        
        course_code = request.POST.get('course', '').strip()
        stream_id = request.POST.get('stream', '').strip()
        try:
            student.course = Course.objects.get(code=course_code)
        except Course.DoesNotExist:
            pass

        try:
            student.stream = Stream.objects.get(id=stream_id) if stream_id else None
        except Stream.DoesNotExist:
            pass
            
        student.user.save()
        student.save()
        messages.success(request, "Student records updated cleanly.")
        return redirect('attendance:manage_students')
        
    courses = Course.objects.all()
    streams = Stream.objects.filter(course=student.course)
    return render(request, 'attendance/edit_student.html', {
        'student': student, 
        'active_courses_list': courses,
        'streams': streams
    })


@login_required
@transaction.atomic
def delete_student(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    student = get_object_or_404(StudentProfile, pk=pk)
    user = student.user
    student.delete()
    user.delete()
    messages.success(request, "Student deleted from database repository.")
    return redirect('attendance:manage_students')


# =========================================================================
# NEW. DEPARTMENTS MANAGEMENT
# =========================================================================

@login_required
@transaction.atomic
def manage_departments(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        name = request.POST.get('department_name', '').strip()
        if name:
            Department.objects.get_or_create(name=name)
            messages.success(request, f"Department '{name}' successfully integrated.")
        else:
            messages.error(request, "Department name value cannot be blank.")
        return redirect('attendance:manage_departments')

    departments = Department.objects.all()
    return render(request, 'attendance/manage_departments.html', {'departments': departments})

@login_required
def add_department(request):
    """
    Handles the structural ingestion and creation of a new academic department.
    """
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        dept_name = request.POST.get('department_name', '').strip()
        
        if not dept_name:
            messages.error(request, "Department name cannot be empty.")
            return redirect('attendance:manage_departments')
            
        # Check for duplication framework arrays
        if Department.objects.filter(name__iexact=dept_name).exists():
            messages.error(request, f"The department '{dept_name}' already exists.")
            return redirect('attendance:manage_departments')
            
        # Create structural track entry
        Department.objects.create(name=dept_name)
        messages.success(request, f"Department '{dept_name}' successfully configured.")
        
    return redirect('attendance:manage_departments')

    
@login_required
@transaction.atomic
def edit_department(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('department_name', '').strip()
        if name:
            department.name = name
            department.save()
            messages.success(request, "Department structural identifier modified successfully.")
            return redirect('attendance:manage_departments')
        else:
            messages.error(request, "Department fields cannot be blank.")
            
    return render(request, 'attendance/edit_department.html', {'department': department})


@login_required
@transaction.atomic
def delete_department(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    department = get_object_or_404(Department, pk=pk)
    name = department.name
    department.delete()
    messages.success(request, f"Department '{name}' cleanly detached from application context.")
    return redirect('attendance:manage_departments')


# =========================================================================
# 3. COURSES MANAGEMENT (UPDATED FOR DEPARTMENTS)
# =========================================================================

@login_required
@transaction.atomic
def manage_courses(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'single' or 'course_code' in request.POST:
            code = (request.POST.get('course_code') or request.POST.get('code', '')).strip()
            name = (request.POST.get('course_name') or request.POST.get('name', '')).strip()
            department_id = request.POST.get('department', '').strip()
            
            if code and name:
                dept = Department.objects.filter(id=department_id).first() if department_id else None
                course, created = Course.objects.get_or_create(code=code, defaults={'name': name, 'department': dept})
                if not created:
                    course.name = name
                    course.department = dept
                    course.save()
                messages.success(request, f"Course Program '{code}' integrated.")
            return redirect('attendance:manage_courses')

        elif action == 'bulk' and request.FILES.get('csv_file'):
            csv_file = request.FILES['csv_file']
            decoded_file = csv_file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(decoded_file), delimiter=',')
            next(reader, None)

            for row in reader:
                if len(row) >= 2:
                    c_code, c_name = row[0].strip(), row[1].strip()
                    if c_code:  # Safety guard for bulk uploads
                        Course.objects.get_or_create(code=c_code, defaults={'name': c_name})
            return redirect('attendance:manage_courses')

    # FIX: Exclude any corrupted empty or null text rows from the stream
    courses = Course.objects.select_related('department').exclude(code="").exclude(code__isnull=True)
    departments = Department.objects.all()
    
    return render(request, 'attendance/manage_courses.html', {
        'courses': courses,
        'departments': departments
    })

@login_required
@transaction.atomic
def edit_course(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        course.code = (request.POST.get('course_code') or request.POST.get('code', '')).strip()
        course.name = (request.POST.get('course_name') or request.POST.get('name', '')).strip()
        department_id = request.POST.get('department', '').strip()
        
        if department_id:
            course.department = Department.objects.filter(id=department_id).first()
        else:
            course.department = None
            
        course.save()
        messages.success(request, "Course program updated successfully.")
        return redirect('attendance:manage_courses')
        
    departments = Department.objects.all()
    return render(request, 'attendance/edit_course.html', {
        'course': course,
        'departments': departments
    })


@login_required
@transaction.atomic
def delete_course(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    course = get_object_or_404(Course, pk=pk)
    course.delete()
    messages.success(request, "Course program completely wiped from scope variables.")
    return redirect('attendance:manage_courses')


# =========================================================================
# 4. COURSE UNITS MANAGEMENT
# =========================================================================
@login_required
@transaction.atomic
def manage_course_units(request):
    print(f"\n--- [DEBUG] Enter manage_course_units View ---")
    print(f"[DEBUG] Request Method: {request.method} | User: {request.user.username} | Role: {getattr(request.user, 'role', 'None')}")

    # 1. Authorization Guard
    if request.user.role != User.IS_ADMIN:
        print(f"[DEBUG] Access Denied: User role {request.user.role} is not IS_ADMIN.")
        return HttpResponse("Unauthorized", status=403)

    # 2. Handle Form Submissions
    if request.method == 'POST':
        action = request.POST.get('action')
        print(f"[DEBUG] POST Action received: '{action}'")
        
        # --- SINGLE COURSE UNIT CREATION / UPDATE ---
        if action == 'single' or 'course_unit_code' in request.POST:
            print(f"[DEBUG] Routing to SINGLE course unit processing")
            code = (request.POST.get('course_unit_code') or request.POST.get('code', '')).strip()
            name = (request.POST.get('course_unit_name') or request.POST.get('name', '')).strip()
            course_code = request.POST.get('course_code', '').strip()
            
            print(f"[DEBUG] Parsed Single Data -> Code: '{code}', Name: '{name}', Parent Course Code: '{course_code}'")

            if code and name and course_code:
                try:
                    course = Course.objects.get(code=course_code)
                    print(f"[DEBUG] Found parent Course instance PK '{course.pk}' for code '{course_code}'")
                    
                    obj, created = CourseUnit.objects.update_or_create(
                        code=code, 
                        defaults={'name': name, 'course': course}
                    )
                    
                    if created:
                        print(f"[DEBUG] SUCCESS: Created NEW CourseUnit PK '{obj.pk}' (Code: {code})")
                        messages.success(request, f"Unit module '{code}' created and linked directly to parent.")
                    else:
                        print(f"[DEBUG] SUCCESS: UPDATED existing CourseUnit PK '{obj.pk}' (Code: {code})")
                        messages.success(request, f"Unit module '{code}' successfully updated.")
                        
                except Course.DoesNotExist:
                    print(f"[DEBUG] ERROR: Parent Course with code '{course_code}' does not exist in database.")
                    messages.error(request, "Failed addition: Specified Parent Course doesn't exist.")
            else:
                print(f"[DEBUG] ERROR: Validation failed. One or more required fields were evaluated as empty strings.")
                messages.error(request, "Failed addition: Missing required fields.")
                
            return redirect('attendance:manage_course_units')

        # --- BULK CSV UPLOAD ---
        elif action == 'bulk' and request.FILES.get('csv_file'):
            csv_file = request.FILES['csv_file']
            print(f"[DEBUG] Routing to BULK CSV processing. File name: '{csv_file.name}'")
            
            decoded_file = csv_file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(decoded_file), delimiter=',')
            next(reader, None)  # Skip CSV header row

            saved_count = 0
            failed_rows = []

            for index, row in enumerate(reader, start=2):
                if len(row) >= 4:
                    c_code = row[0].strip()
                    cu_code = row[2].strip()
                    cu_name = row[3].strip()
                    
                    if not c_code or not cu_code or not cu_name:
                        failed_rows.append(f"Row {index} (Missing fields)")
                        continue

                    try:
                        course = Course.objects.get(code=c_code)
                        CourseUnit.objects.update_or_create(
                            code=cu_code, 
                            defaults={'name': cu_name, 'course': course}
                        )
                        saved_count += 1
                    except Course.DoesNotExist:
                        failed_rows.append(f"Row {index} (Course code '{c_code}' not found)")
                        continue
                else:
                    failed_rows.append(f"Row {index} (Malformed column structure)")

            if saved_count > 0:
                messages.success(request, f"Successfully processed {saved_count} course unit(s).")
            if failed_rows:
                messages.error(request, f"Skipped rows due to errors: {', '.join(failed_rows)}")

            return redirect('attendance:manage_course_units')

    # 3. Handle GET Request / Fetch View Context
    course_units = CourseUnit.objects.select_related('course').all()
    courses = Course.objects.all()
    
    for c in courses:
        c.course_code = c.code
        c.course_name = c.name

    return render(request, 'attendance/manage_course_units.html', {
        'course_units': course_units, 
        'courses': courses,
        'active_courses_list': courses
    })



@login_required
@transaction.atomic
def edit_course_unit(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
        
    unit = get_object_or_404(CourseUnit, pk=pk)
    if request.method == 'POST':
        unit.code = (request.POST.get('course_unit_code') or request.POST.get('code', '')).strip()
        unit.name = (request.POST.get('course_unit_name') or request.POST.get('name', '')).strip()
        
        course_code = request.POST.get('course_code', '').strip()
        try:
            unit.course = Course.objects.get(code=course_code)
        except Course.DoesNotExist:
            pass
            
        unit.save()
        messages.success(request, "Subject Module context definitions modified successfully.")
        return redirect('attendance:manage_course_units')
        
    courses = Course.objects.all()
    return render(request, 'attendance/edit_course_unit.html', {'unit': unit, 'active_courses_list': courses})


@login_required
@transaction.atomic
def delete_course_unit(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
        
    unit = get_object_or_404(CourseUnit, pk=pk)
    unit.delete()
    messages.success(request, "Module index row cleanly detached and purged.")
    return redirect('attendance:manage_course_units')


# =========================================================================
# 5. CORE ADMINISTRATIVE DASHBOARD & UTILITIES
# =========================================================================

@login_required
def admin_dashboard(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    total_courses = Course.objects.count()
    total_teachers = TeacherProfile.objects.count()
    total_students = StudentProfile.objects.count()
    total_present = AttendanceRecord.objects.filter(status='PRESENT').count()
    total_records = AttendanceRecord.objects.count()
    overall_rate = round((total_present / total_records) * 100, 2) if total_records > 0 else 0

    teacher_stats = []
    teachers = TeacherProfile.objects.all()
    for t in teachers:
        present = AttendanceRecord.objects.filter(
            session__timetable_entry__teacher=t, status='PRESENT'
        ).count()
        total = AttendanceRecord.objects.filter(session__timetable_entry__teacher=t).count()
        rate = round((present / total) * 100, 2) if total > 0 else 0
        teacher_stats.append({'name': t.name, 'rate': rate})
    
    teacher_names = [ts['name'] for ts in teacher_stats]
    teacher_rates = [ts['rate'] for ts in teacher_stats]

    student_stats = StudentProfile.objects.annotate(
        present=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT')),
        total=Count('attendancerecord')
    ).order_by('-total')[:10]
    
    student_labels = [s.name for s in student_stats]
    student_present = [s.present for s in student_stats]
    student_absent = [s.total - s.present for s in student_stats]

    stream_counts = Stream.objects.select_related('course').annotate(
        student_count=Count('students')
    ).order_by('name')

    context = {
        'total_courses': total_courses,
        'total_teachers': total_teachers,
        'total_students': total_students,
        'overall_rate': overall_rate,
        'teacher_names': json.dumps(teacher_names),
        'teacher_rates': json.dumps(teacher_rates),
        'student_labels': json.dumps(student_labels),
        'student_present': json.dumps(student_present),
        'student_absent': json.dumps(student_absent),
        'stream_counts': stream_counts,
    }
    return render(request, 'attendance/admin_dashboard.html', context)


@login_required
@transaction.atomic
def bulk_upload_courses(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        decoded_file = csv_file.read().decode('utf-8')
        reader = csv.reader(io.StringIO(decoded_file), delimiter=',')
        next(reader, None)

        for row in reader:
            if len(row) >= 4:
                c_code, c_name, cu_code, cu_name = row[0].strip(), row[1].strip(), row[2].strip(), row[3].strip()
                course, _ = Course.objects.get_or_create(code=c_code, defaults={'name': c_name})
                CourseUnit.objects.get_or_create(code=cu_code, defaults={'name': cu_name, 'course': course})
        return redirect('attendance:admin_dashboard')
    return render(request, 'attendance/upload_courses.html')


@login_required
@transaction.atomic
def bulk_upload_teachers(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        decoded_file = csv_file.read().decode('utf-8')
        reader = csv.reader(io.StringIO(decoded_file), delimiter=',')
        next(reader, None)

        for row in reader:
            if len(row) >= 2:
                name, email = row[0].strip(), row[1].strip()
                username = email.split('@')[0]
                password = generate_secure_password()

                user, created = User.objects.get_or_create(email=email, defaults={
                    'username': username,
                    'role': User.IS_TEACHER,
                    'raw_password_archive': password
                })
                if created:
                    user.set_password(password)
                    user.save()
                    TeacherProfile.objects.create(user=user, name=name)
        return redirect('attendance:export_credentials', role_type='teachers')
    return render(request, 'attendance/upload_teachers.html')


@login_required
@transaction.atomic
def bulk_upload_students(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        decoded_file = csv_file.read().decode('utf-8')
        reader = csv.reader(io.StringIO(decoded_file), delimiter=',')
        next(reader, None)

        for row in reader:
            if len(row) >= 5:
                name, reg_num, email, course_code, stream_name = row[0].strip(), row[1].strip(), row[2].strip(), row[3].strip(), row[4].strip()
                try:
                    course = Course.objects.get(code=course_code)
                    
                    from attendance.models import Stream
                    stream_obj, _ = Stream.objects.get_or_create(name=stream_name, course=course)
                    
                    password = generate_secure_password()
                    user, created = User.objects.get_or_create(email=email, defaults={
                        'username': reg_num.replace('/', '_'),
                        'role': User.IS_STUDENT,
                        'raw_password_archive': password
                    })
                    if created:
                        user.set_password(password)
                        user.save()
                        StudentProfile.objects.create(user=user, reg_number=reg_num, name=name, course=course, stream=stream_obj)
                except Course.DoesNotExist:
                    continue
        return redirect('attendance:export_credentials', role_type='students')
    return render(request, 'attendance/upload_students.html')

import json
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.contrib import messages
from datetime import datetime
from .models import TimetableBatch, TimetableEntry, CourseUnit, TeacherProfile, Stream
import json
import json
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.contrib import messages
from datetime import datetime
from .models import TimetableBatch, TimetableEntry, CourseUnit, TeacherProfile, Stream

import json
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from attendance.models import User, Stream, TimetableBatch, TimetableEntry, CourseUnit, TeacherProfile


@login_required
def manage_timetable(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # Fetch active batch
    batch = TimetableBatch.objects.filter(is_active=True, is_revoked=False).order_by('-uploaded_at').first()
    
    # Handle optional clear action to wipe old test data and start fresh
    if request.GET.get('action') == 'clear_all' and batch:
        TimetableEntry.objects.filter(batch=batch).delete()
        messages.success(request, "All old timetable entries have been cleared. You now have a clean slate.")
        return redirect('attendance:manage_timetable')

    streams = Stream.objects.select_related('course').all()
    
    # Strictly filter and count entries that actually belong to each stream for this batch
    streams_list = []
    for stream in streams:
        entry_count = 0
        if batch:
            entry_count = TimetableEntry.objects.filter(batch=batch, stream=stream).count()
            
        streams_list.append({
            'id': stream.id,
            'name': stream.name,
            'course_name': stream.course.name if stream.course else "No Course Assigned",
            'has_timetable': entry_count > 0,  # True only if entries exist specifically for this class
            'entry_count': entry_count
        })

    return render(request, 'attendance/manage_timetable_list.html', {
        'streams_list': streams_list,
        'batch': batch
    })

@login_required
@transaction.atomic
def upload_timetable(request, stream_id):
    """
    Interactive Timetable Matrix Editor focused strictly on an individual stream/class.
    Omits stream selection from cells and saves all configurations under the current stream.
    """
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    stream_obj = get_object_or_404(Stream, id=stream_id)

    # Retrieve or initialize active batch configuration
    batch = TimetableBatch.objects.filter(is_active=True, is_revoked=False).order_by('-uploaded_at').first()
    if not batch:
        batch = TimetableBatch.objects.create(week_start_date=datetime.now().date(), is_active=True)

    DAYS = [code for code, _ in TimetableEntry.DAYS_OF_WEEK]

    if request.method == 'POST':
        # Wipe existing records for this stream under the current batch to overwrite cleanly
        TimetableEntry.objects.filter(batch=batch, stream=stream_obj).delete()

        row_indexes = request.POST.getlist('row_index')
        
        for idx in row_indexes:
            start_time = request.POST.get(f'start_time_{idx}')
            end_time = request.POST.get(f'end_time_{idx}')
            
            if not start_time or not end_time:
                continue

            for day_code in DAYS:
                cu_code = request.POST.get(f'cu_{idx}_{day_code}')
                teacher_id = request.POST.get(f'teacher_{idx}_{day_code}')

                # Save record if a course unit is chosen for this slot
                if cu_code:
                    try:
                        course_unit = CourseUnit.objects.get(code=cu_code)
                        teacher = TeacherProfile.objects.get(id=teacher_id) if teacher_id else None
                        
                        if teacher:
                            TimetableEntry.objects.create(
                                batch=batch,
                                day=day_code,
                                start_time=start_time,
                                end_time=end_time,
                                course_unit=course_unit,
                                teacher=teacher,
                                stream=stream_obj # Contextually assigned class stream
                            )
                    except Exception as e:
                        pass

        messages.success(request, f"Timetable configuration saved successfully for class '{stream_obj.name}'.")
        return redirect('attendance:upload_timetable', stream_id=stream_obj.id)

    # GET Process: Construct matrix rows specifically for this stream
    existing_entries = TimetableEntry.objects.filter(batch=batch, stream=stream_obj).select_related('course_unit', 'teacher')
    
    # Group entries by unique time slots
    time_slots_map = {}
    for entry in existing_entries:
        key = (entry.start_time.strftime('%H:%M'), entry.end_time.strftime('%H:%M'))
        if key not in time_slots_map:
            time_slots_map[key] = {}
        time_slots_map[key][entry.day] = entry

    matrix_rows = []
    counter = 0
    for (start, end), day_entries in time_slots_map.items():
        slots = []
        for day_code in DAYS:
            slots.append({
                'day_code': day_code,
                'entry': day_entries.get(day_code)
            })
        matrix_rows.append({
            'index': counter,
            'start': start,
            'end': end,
            'slots': slots
        })
        counter += 1

    # Generate qualified instructor mapping arrays
    cu_lecturer_map = {}
    for cu in CourseUnit.objects.all():
        teachers = TeacherProfile.objects.filter(courses=cu.course)
        if not teachers.exists():
            teachers = TeacherProfile.objects.all()
        cu_lecturer_map[str(cu.code)] = [
            {'id': t.id, 'name': t.name} for t in teachers
        ]

    context = {
        'stream': stream_obj,
        'matrix_rows': matrix_rows,
        'days': TimetableEntry.DAYS_OF_WEEK,
        'course_units': CourseUnit.objects.all(),
        'cu_lecturer_map_json': json.dumps(cu_lecturer_map),
    }
    return render(request, 'attendance/upload_timetable.html', context)

def download_template(request, template_type):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{template_type}_template.csv"'

    if template_type == 'courses':
        content = "course_code,course_name,course_unit_code,course_unit_name\nC001,Computer Science,CS101,Programming Basics\nC001,Computer Science,CS102,Data Structures"
    elif template_type == 'teachers':
        content = "teachers_name,email\nJohn Doe,john@example.com"
    elif template_type == 'students':
        content = "students_name,registration_number,email,course,stream_name\nAlice Smith,2024/001,alice@example.com,CIT,Year 1 CIT A"
    elif template_type == 'timetable':
        content = "day,start_time,end_time,course_unit_code,teacher_email,stream_name\nMON,08:30,10:00,CIT101,james.muwonge@utc.ac.ug,Year 1 CIT A"
    else:
        content = ""

    response.write(content)
    return response


# =========================================================================
# 6. CREDENTIAL EXPORTS ENGINE (EXCEL OUTPUT FORMAT)
# =========================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

@login_required
def export_credentials(request, role_type):
    wb = Workbook()
    ws = wb.active
    ws.title = "Credentials"

    headers = ['Identifier/Email', 'Name', 'Generated Password']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')

    row_num = 2
    if role_type == 'teachers':
        profiles = TeacherProfile.objects.select_related('user').all()
        for p in profiles:
            ws.cell(row=row_num, column=1, value=p.user.email)
            ws.cell(row=row_num, column=2, value=p.name)
            ws.cell(row=row_num, column=3, value=p.user.raw_password_archive)
            row_num += 1
    elif role_type == 'students':
        profiles = StudentProfile.objects.select_related('user').all()
        for p in profiles:
            ws.cell(row=row_num, column=1, value=p.reg_number)
            ws.cell(row=row_num, column=2, value=p.name)
            ws.cell(row=row_num, column=3, value=p.user.raw_password_archive)
            row_num += 1

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        column = col[0].column_letter
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{role_type}_credentials.xlsx"'
    wb.save(response)
    return response


# =========================================================================
# NEW. ADMINISTRATIVE REPORTS VIEW
# =========================================================================


@login_required
def admin_report_page(request):
    # Enforce Admin-only access
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
        
    departments = Department.objects.all()
    
    # Extract structural query constraint parameter directly from backend request arrays
    selected_dept_id = request.GET.get('filter_dept')
    
    # Start with base layout mapping matching all active structural course matrices
    courses = Course.objects.select_related('department').all()
    
    # Execute structural filtering directly inside database processing layers if constraint exists
    if selected_dept_id:
        courses = courses.filter(department_id=selected_dept_id)
        try:
            selected_dept_id = int(selected_dept_id)  # Standardize type matching for UI template rules
        except ValueError:
            selected_dept_id = None

    return render(request, 'attendance/admin_report.html', {
        'departments': departments,
        'courses': courses,
        'selected_dept_id': selected_dept_id
    })

'''
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
import json
from attendance.models import User, Stream, AttendanceRecord, AttendanceSession, StudentProfile

@login_required
def analytics_dashboard(request):

    # Enforce strict Admin-only access rule matrix
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # Extract optional filter parameters from template request scopes
    selected_stream_id = request.GET.get('stream')
    
    # Base query sets targeting core transactional datasets
    records = AttendanceRecord.objects.all()
    sessions = AttendanceSession.objects.select_related(
        'timetable_entry__teacher__user', 
        'timetable_entry__stream'
    ).order_by('-date_marked')
    
    # Dynamic application of scoping filters if requested by the end-user
    if selected_stream_id and selected_stream_id != 'all':
        records = records.filter(student__stream_id=selected_stream_id)
        sessions = sessions.filter(timetable_entry__stream_id=selected_stream_id)

    # 1. Compute Live Summary Statistics Metrics Matrix
    total_records = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()

    stats = {
        'total_records': total_records,
        'present_rate': round((present_count / total_records) * 100, 1) if total_records > 0 else 0,
        'absent_rate': round((absent_count / total_records) * 100, 1) if total_records > 0 else 0,
        'late_rate': 0  # Maintained at 0 since model STATUS_CHOICES only specifies PRESENT/ABSENT
    }

    # 2. Compile Lecturer Location Audit Log Dataset Dynamically
    audit_logs = []
    # Fetching the top 20 latest session instances to prevent UI buffer layout breaks
    for session in sessions[:20]:
        t_entry = session.timetable_entry
        has_coords = session.teacher_latitude is not None and session.teacher_longitude is not None
        
        location_str = f"{session.teacher_latitude}, {session.teacher_longitude}" if has_coords else "Not captured"
        
        audit_logs.append({
            "date": session.date_marked.strftime("%Y-%m-%d") if session.date_marked else "—",
            "class": t_entry.stream.name if t_entry.stream else "—",
            "lecturer": t_entry.teacher.user.email if t_entry.teacher and t_entry.teacher.user else "—",
            "location": location_str,
            "ip": "—",  # Placeholder asset as IP capture hooks are not defined inside current models schema
            "verified": has_coords  # Flag verified true strictly if geolocation values exist
        })

    # 3. Pull Top Students Metrics Arrays via Database Annotations
    students = StudentProfile.objects.select_related('stream').annotate(
        total_rec=Count('attendancerecord'),
        present_rec=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT'))
    ).filter(total_rec__gt=0)
    
    top_students_raw = []
    for s in students:
        rate = (s.present_rec / s.total_rec) * 100
        top_students_raw.append({
            "name": s.name,
            "reg": s.reg_number,
            "class": s.stream.name if s.stream else "—",
            "rate_num": rate,
            "rate": f"{round(rate, 1)}%"
        })
    
    # Sort and slice dataset to grab top performing arrays
    top_students = sorted(top_students_raw, key=lambda x: x['rate_num'], reverse=True)[:5]

    # 4. Generate Aggregated Payload Engines for ChartJS
    # Chart A: Global Present vs Absent distribution count vectors
    chart_dist_data = [present_count, absent_count]
    
    # Chart B: Generate stream breakdown labels and cross-cutting attendance percentage matrices
    streams_data = Stream.objects.annotate(
        total=Count('students__attendancerecord'),
        present=Count('students__attendancerecord', filter=Q(students__attendancerecord__status='PRESENT'))
    ).filter(total__gt=0)
    
    stream_labels = [stream.name for stream in streams_data]
    stream_rates = [round((stream.present / stream.total) * 100, 1) for stream in streams_data]

    # Context compilation pass payload arrays safely out into frontend DOM nodes
    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_students': top_students,
        'streams': Stream.objects.all(),
        'selected_stream': selected_stream_id,
        'chart_dist_data': json.dumps(chart_dist_data),
        'stream_labels': json.dumps(stream_labels),
        'stream_rates': json.dumps(stream_rates),
    }
    return render(request, 'attendance/analytics_dashboard.html', context)
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # 1. Summary Statistics Metrics
    stats = {
        'total_records': 10,
        'present_rate': 90,
        'absent_rate': 10,
        'late_rate': 0
    }

    # 2. Lecturer Location Audit Log Dataset
    audit_logs = [
        {"date": "2026-06-26 @ 15:21", "class": "NDEE Weekend", "lecturer": "laban.omenyo.ai@gmail.com", "location": "0.3375, 32.5943 (±193642m)", "ip": "41.210.154.15", "verified": True},
        {"date": "2026-06-25 @ 21:45", "class": "NDME A", "lecturer": "technologiestescal@gmail.com", "location": "-0.6195, 30.6697 (±65m)", "ip": "41.210.154.139", "verified": True},
        {"date": "2026-05-18 @ 16:20", "class": "NDEE Weekend", "lecturer": "technologiestescal@gmail.com", "location": "-0.5470, 30.2457 (±130m)", "ip": "154.72.206.194", "verified": True},
        {"date": "2026-05-10 @ 09:13", "class": "NDEE Weekend", "lecturer": "technologiestescal@gmail.com", "location": "-0.5469, 30.2457 (±68m)", "ip": "41.210.146.63", "verified": True},
        {"date": "2026-05-10 @ 22:06", "class": "NDME A", "lecturer": "technologiestescal@gmail.com", "location": "Not captured", "ip": "41.210.146.63", "verified": True},
        {"date": "2026-05-10 @ 21:54", "class": "NDME A", "lecturer": "technologiestescal@gmail.com", "location": "-0.5502, 30.2441 (±30m)", "ip": "41.210.146.63", "verified": False},
        {"date": "2026-05-10 @ 12:36", "class": "NDA A", "lecturer": "technologiestescal@gmail.com", "location": "-0.5469, 30.2457 (±61m)", "ip": "154.72.206.194", "verified": True},
        {"date": "2026-05-10 @ 12:35", "class": "NDA A", "lecturer": "technologiestescal@gmail.com", "location": "Not captured", "ip": "154.72.206.194", "verified": False},
        {"date": "2026-05-08 @ 22:54", "class": "NDEE Weekend", "lecturer": "—", "location": "Not captured", "ip": "—", "verified": False},
        {"date": "2026-05-09 @ 22:33", "class": "NDA A", "lecturer": "—", "location": "Not captured", "ip": "—", "verified": False},
        {"date": "2026-05-09 @ 10:57", "class": "NDEE Weekend", "lecturer": "—", "location": "Not captured", "ip": "—", "verified": False},
        {"date": "2026-05-09 @ 10:39", "class": "NDME A", "lecturer": "—", "location": "Not captured", "ip": "—", "verified": False},
    ]

    # 3. Top Students Placeholder Data Structure
    top_students = [
        {"name": "Mukasa John", "reg": "2025/NDEE/021", "class": "NDEE Weekend", "rate": "98%"},
        {"name": "Ainomugisha Dianah", "reg": "2025/NDME/104", "class": "NDME A", "rate": "96%"},
        {"name": "Mwesigye Brian", "reg": "2025/NDA/008", "class": "NDA A", "rate": "95%"},
    ]

    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_students': top_students,
    }
    return render(request, 'attendance/analytics_dashboard.html', context)



'''
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from datetime import datetime, timedelta
import json
import csv
from attendance.models import User, Department, Course, Stream, AttendanceRecord, AttendanceSession, StudentProfile

import csv
import json
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.core.exceptions import FieldError

# Assuming your models are imported like this:
# from .models import User, Department, Course, Stream, StudentProfile, AttendanceRecord, AttendanceSession

@login_required
def analytics_dashboard(request):
    # Enforce strict Admin-only access rule matrix
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # 1. CAPTURE DETAILED REPORT FILTERS
    filter_dept = request.GET.get('filter_dept')
    filter_course = request.GET.get('filter_course')
    filter_stream = request.GET.get('filter_stream')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    quick_range = request.GET.get('quick_range')

    # Base entity dropdown datasets supporting cascading display lists
    all_departments = Department.objects.all()
    selectable_courses = Course.objects.all()
    selectable_streams = Stream.objects.all()

    if filter_dept:
        selectable_courses = selectable_courses.filter(department_id=filter_dept)
    if filter_course:
        selectable_streams = selectable_streams.filter(course__code=filter_course)

    # 2. RESOLVE TIME DATA WINDOW CONSTRAINTS
    today = datetime.now().date()
    start_date = None
    end_date = None

    if quick_range == 'today':
        start_date = today
        end_date = today
    elif quick_range == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif quick_range == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    else:
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

    # Build cross-cutting performance filter arrays
    record_date_filter = Q()
    if start_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__gte=start_date)
    if end_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__lte=end_date)

    # 3. COMPILE COMPLETE DETAILED STUDENT POPULATION ROSTER
    report_students_queryset = StudentProfile.objects.select_related('stream', 'course__department').all()
    
    if filter_stream:
        report_students_queryset = report_students_queryset.filter(stream_id=filter_stream)
    elif filter_course:
        report_students_queryset = report_students_queryset.filter(course__code=filter_course)
    elif filter_dept:
        report_students_queryset = report_students_queryset.filter(course__department_id=filter_dept)

    # Calculate absolute percentages over selected date parameters
    annotated_students = report_students_queryset.annotate(
        total_filtered_records=Count('attendancerecord', filter=record_date_filter),
        present_filtered_records=Count('attendancerecord', filter=record_date_filter & Q(attendancerecord__status='PRESENT'))
    ).order_by('name')

    detailed_student_reports = []
    for s in annotated_students:
        tot = s.total_filtered_records
        pres = s.present_filtered_records
        percentage_val = round((pres / tot) * 100, 1) if tot > 0 else 0.0
        
        detailed_student_reports.append({
            "name": s.name,
            "reg_number": s.reg_number,
            "stream_name": s.stream.name if s.stream else "—",
            "total_logs": tot,
            "present_logs": pres,
            "attendance_percentage": f"{percentage_val}%",
            "percentage_num": percentage_val
        })

    # 4. INTERCEPT SPECIFIC DETAILED EXPORT REQUESTS
    if request.GET.get('export') == 'detailed_student_csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="detailed_student_report_{today}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Registration Number', 'Assigned Stream', 'Total Sessions Checked', 'Sessions Attended', 'Attendance Rate'])
        for item in detailed_student_reports:
            writer.writerow([item['name'], item['reg_number'], item['stream_name'], item['total_logs'], item['present_logs'], item['attendance_percentage']])
        return response

    # 5. PRESERVE ORIGINAL ANALYTICS COMPONENT ARRAYS
    selected_stream_id = request.GET.get('stream')
    records = AttendanceRecord.objects.all()
    sessions = AttendanceSession.objects.select_related('timetable_entry__teacher__user', 'timetable_entry__stream').order_by('-date_marked')
    
    if selected_stream_id and selected_stream_id != 'all':
        records = records.filter(student__stream_id=selected_stream_id)
        sessions = sessions.filter(timetable_entry__stream_id=selected_stream_id)

    # 6. INTERCEPT GLOBAL CSV EXPORT REQUEST
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Reg Number', 'Class/Stream', 'Status', 'Date'])
        for record in records.select_related('student__stream', 'session'):
            date_val = record.session.date_marked.strftime('%Y-%m-%d') if hasattr(record, 'session') and record.session.date_marked else 'N/A'
            stream_name = record.student.stream.name if record.student.stream else 'N/A'
            writer.writerow([record.student.name, record.student.reg_number, stream_name, record.status, date_val])
        return response

    # Process original global metrics
    total_records = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()

    stats = {
        'total_records': total_records,
        'present_rate': round((present_count / total_records) * 100, 1) if total_records > 0 else 0,
        'absent_rate': round((absent_count / total_records) * 100, 1) if total_records > 0 else 0,
    }

    # Location Log Parsing
    audit_logs = []
    for session in sessions[:20]:
        t_entry = session.timetable_entry
        has_coords = session.teacher_latitude is not None and session.teacher_longitude is not None
        location_str = f"{session.teacher_latitude}, {session.teacher_longitude}" if has_coords else "Not captured"
        audit_logs.append({
            "date": session.date_marked.strftime("%Y-%m-%d") if session.date_marked else "—",
            "class": t_entry.stream.name if t_entry.stream else "—",
            "lecturer": t_entry.teacher.user.email if t_entry.teacher and t_entry.teacher.user else "—",
            "location": location_str,
            "ip": "—",  
            "verified": has_coords  
        })

    # Standard Top-20 Matrix Ingestion 
    students = StudentProfile.objects.select_related('stream').annotate(
        total_rec=Count('attendancerecord'),
        present_rec=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT')),
        absent_rec=Count('attendancerecord', filter=Q(attendancerecord__status='ABSENT'))
    ).filter(total_rec__gt=0)
    
    students_raw = []
    for s in students:
        p_rate = (s.present_rec / s.total_rec) * 100
        a_rate = (s.absent_rec / s.total_rec) * 100
        students_raw.append({
            "name": s.name,
            "reg": s.reg_number,
            "class": s.stream.name if s.stream else "—",
            "p_rate_num": p_rate,
            "a_rate_num": a_rate,
            "p_rate": f"{round(p_rate, 1)}%",
            "a_rate": f"{round(a_rate, 1)}%"
        })
    
    top_present_students = sorted(students_raw, key=lambda x: x['p_rate_num'], reverse=True)[:20]
    top_absent_students = sorted(students_raw, key=lambda x: x['a_rate_num'], reverse=True)[:20]

    # 7. ADDED/UPDATED: Lecturer Attendance Submission Compliance Report Scope Filtering
    TimetableEntry = AttendanceSession._meta.get_field('timetable_entry').related_model
    
    compliance_sessions = AttendanceSession.objects.all()
    timetable_entries = TimetableEntry.objects.select_related('teacher__user').all()
    
    # Apply structural Department, Course, and Stream scope constraints dynamically
    if filter_stream:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream_id=filter_stream)
        timetable_entries = timetable_entries.filter(stream_id=filter_stream)
    elif filter_course:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream__course__code=filter_course)
        timetable_entries = timetable_entries.filter(stream__course__code=filter_course)
    elif filter_dept:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream__course__department_id=filter_dept)
        timetable_entries = timetable_entries.filter(stream__course__department_id=filter_dept)

    if selected_stream_id and selected_stream_id != 'all':
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream_id=selected_stream_id)
        timetable_entries = timetable_entries.filter(stream_id=selected_stream_id)

    total_system_sessions = compliance_sessions.count()
    teacher_counts = compliance_sessions.values('timetable_entry__teacher_id').annotate(count=Count('id'))
    counts_map = {item['timetable_entry__teacher_id']: item['count'] for item in teacher_counts if item['timetable_entry__teacher_id']}

    teachers_raw = []
    seen_teachers = set()
    for entry in timetable_entries:
        if entry.teacher and entry.teacher.id not in seen_teachers:
            seen_teachers.add(entry.teacher.id)
            email = entry.teacher.user.email if entry.teacher.user else "—"
            submitted = counts_map.get(entry.teacher.id, 0)
            rate = round((submitted / total_system_sessions) * 100, 1) if total_system_sessions > 0 else 0
            teachers_raw.append({'email': email, 'submitted': submitted, 'rate': f"{rate}%"})

    top_submitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=True)[:20]
    top_unsubmitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=False)[:20]

    # Chart Generation Data Processing
    chart_dist_data = [present_count, absent_count]
    streams_data = Stream.objects.annotate(
        total=Count('students__attendancerecord'),
        present=Count('students__attendancerecord', filter=Q(students__attendancerecord__status='PRESENT'))
    ).filter(total__gt=0)
    
    stream_labels = [stream.name for stream in streams_data]
    stream_rates = [round((stream.present / stream.total) * 100, 1) for stream in streams_data]

    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_present_students': top_present_students,
        'top_absent_students': top_absent_students,
        'top_submitted_teachers': top_submitted_teachers,
        'top_unsubmitted_teachers': top_unsubmitted_teachers,
        'streams': Stream.objects.all(),
        'selected_stream': selected_stream_id,
        'chart_dist_data': json.dumps(chart_dist_data),
        'stream_labels': json.dumps(stream_labels),
        'stream_rates': json.dumps(stream_rates),
        
        # Detailed reporting fields context
        'departments': all_departments,
        'courses': selectable_courses,
        'selectable_streams': selectable_streams,
        'detailed_student_reports': detailed_student_reports,
        'filter_dept': filter_dept,
        'filter_course': filter_course,
        'filter_stream': filter_stream,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'quick_range': quick_range,
    }
    return render(request, 'attendance/analytics_dashboard.html', context)
    # Enforce strict Admin-only access rule matrix
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # 1. CAPTURE DETAILED REPORT FILTERS
    filter_dept = request.GET.get('filter_dept')
    filter_course = request.GET.get('filter_course')
    filter_stream = request.GET.get('filter_stream')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    quick_range = request.GET.get('quick_range')

    # Base entity dropdown datasets supporting cascading display lists
    all_departments = Department.objects.all()
    selectable_courses = Course.objects.all()
    selectable_streams = Stream.objects.all()

    if filter_dept:
        selectable_courses = selectable_courses.filter(department_id=filter_dept)
    if filter_course:
        selectable_streams = selectable_streams.filter(course__code=filter_course)

    # 2. RESOLVE TIME DATA WINDOW CONSTRAINTS
    today = datetime.now().date()
    start_date = None
    end_date = None

    if quick_range == 'today':
        start_date = today
        end_date = today
    elif quick_range == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif quick_range == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    else:
        if start_date_str and start_date_str != 'None':
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date_str and end_date_str != 'None':
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

    # Build cross-cutting performance filter arrays
    record_date_filter = Q()
    if start_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__gte=start_date)
    if end_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__lte=end_date)

    # 3. COMPILE COMPLETE DETAILED STUDENT POPULATION ROSTER
    report_students_queryset = StudentProfile.objects.select_related('stream', 'course__department').all()
    
    if filter_stream:
        report_students_queryset = report_students_queryset.filter(stream_id=filter_stream)
    elif filter_course:
        report_students_queryset = report_students_queryset.filter(course__code=filter_course)
    elif filter_dept:
        report_students_queryset = report_students_queryset.filter(course__department_id=filter_dept)

    # Calculate absolute percentages over selected date parameters
    annotated_students = report_students_queryset.annotate(
        total_filtered_records=Count('attendancerecord', filter=record_date_filter),
        present_filtered_records=Count('attendancerecord', filter=record_date_filter & Q(attendancerecord__status='PRESENT'))
    ).order_by('name')

    detailed_student_reports = []
    for s in annotated_students:
        tot = s.total_filtered_records
        pres = s.present_filtered_records
        percentage_val = round((pres / tot) * 100, 1) if tot > 0 else 0.0
        
        detailed_student_reports.append({
            "name": s.name,
            "reg_number": s.reg_number,
            "stream_name": s.stream.name if s.stream else "—",
            "total_logs": tot,
            "present_logs": pres,
            "attendance_percentage": f"{percentage_val}%",
            "percentage_num": percentage_val
        })

    # 4. INTERCEPT SPECIFIC DETAILED EXPORT REQUESTS
    if request.GET.get('export') == 'detailed_student_csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="detailed_student_report_{today}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Registration Number', 'Assigned Stream', 'Total Sessions Checked', 'Sessions Attended', 'Attendance Rate'])
        for item in detailed_student_reports:
            writer.writerow([item['name'], item['reg_number'], item['stream_name'], item['total_logs'], item['present_logs'], item['attendance_percentage']])
        return response

    # 5. PRESERVE ORIGINAL ANALYTICS COMPONENT ARRAYS
    selected_stream_id = request.GET.get('stream')
    records = AttendanceRecord.objects.all()
    sessions = AttendanceSession.objects.select_related('timetable_entry__teacher__user', 'timetable_entry__stream').order_by('-date_marked')
    
    if selected_stream_id and selected_stream_id != 'all':
        records = records.filter(student__stream_id=selected_stream_id)
        sessions = sessions.filter(timetable_entry__stream_id=selected_stream_id)

    # 6. INTERCEPT GLOBAL CSV EXPORT REQUEST
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Reg Number', 'Class/Stream', 'Status', 'Date'])
        for record in records.select_related('student__stream', 'session'):
            date_val = record.session.date_marked.strftime('%Y-%m-%d') if hasattr(record, 'session') and record.session.date_marked else 'N/A'
            stream_name = record.student.stream.name if record.student.stream else 'N/A'
            writer.writerow([record.student.name, record.student.reg_number, stream_name, record.status, date_val])
        return response

    # Process original global metrics
    total_records = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()

    stats = {
        'total_records': total_records,
        'present_rate': round((present_count / total_records) * 100, 1) if total_records > 0 else 0,
        'absent_rate': round((absent_count / total_records) * 100, 1) if total_records > 0 else 0,
    }

    # Location Log Parsing
    audit_logs = []
    for session in sessions[:20]:
        t_entry = session.timetable_entry
        has_coords = session.teacher_latitude is not None and session.teacher_longitude is not None
        location_str = f"{session.teacher_latitude}, {session.teacher_longitude}" if has_coords else "Not captured"
        audit_logs.append({
            "date": session.date_marked.strftime("%Y-%m-%d") if session.date_marked else "—",
            "class": t_entry.stream.name if t_entry.stream else "—",
            "lecturer": t_entry.teacher.user.email if t_entry.teacher and t_entry.teacher.user else "—",
            "location": location_str,
            "ip": "—",  
            "verified": has_coords  
        })

    # Standard Top-20 Matrix Ingestion 
    students = StudentProfile.objects.select_related('stream').annotate(
        total_rec=Count('attendancerecord'),
        present_rec=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT')),
        absent_rec=Count('attendancerecord', filter=Q(attendancerecord__status='ABSENT'))
    ).filter(total_rec__gt=0)
    
    students_raw = []
    for s in students:
        p_rate = (s.present_rec / s.total_rec) * 100
        a_rate = (s.absent_rec / s.total_rec) * 100
        students_raw.append({
            "name": s.name,
            "reg": s.reg_number,
            "class": s.stream.name if s.stream else "—",
            "p_rate_num": p_rate,
            "a_rate_num": a_rate,
            "p_rate": f"{round(p_rate, 1)}%",
            "a_rate": f"{round(a_rate, 1)}%"
        })
    
    top_present_students = sorted(students_raw, key=lambda x: x['p_rate_num'], reverse=True)[:20]
    top_absent_students = sorted(students_raw, key=lambda x: x['a_rate_num'], reverse=True)[:20]

    # 7. RESILIENT CASCADING LOGIC FOR LECTURER COMPLIANCE REPORT
    TimetableEntry = AttendanceSession._meta.get_field('timetable_entry').related_model
    
    compliance_sessions = AttendanceSession.objects.all()
    timetable_entries = TimetableEntry.objects.select_related('teacher__user').all()
    
    # Apply dynamic scoping constraints based on user dropdown selections
    if filter_stream:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream_id=filter_stream)
        timetable_entries = timetable_entries.filter(stream_id=filter_stream)
    elif filter_course:
        try:
            # Look up via standard Stream -> Course relation path
            compliance_sessions = compliance_sessions.filter(timetable_entry__stream__course__code=filter_course)
            timetable_entries = timetable_entries.filter(stream__course__code=filter_course)
        except Exception:
            try:
                # Look up via direct Course schema field link (if any)
                compliance_sessions = compliance_sessions.filter(timetable_entry__course__code=filter_course)
                timetable_entries = timetable_entries.filter(course__code=filter_course)
            except Exception:
                pass
    elif filter_dept:
        try:
            # Tier 1 Lookup: Query via Stream -> Course -> Department link mapping
            compliance_sessions = compliance_sessions.filter(timetable_entry__stream__course__department_id=filter_dept)
            timetable_entries = timetable_entries.filter(stream__course__department_id=filter_dept)
            
            if not timetable_entries.exists():
                raise FieldError()
        except (FieldError, Exception):
            try:
                # Tier 2 Lookup: Query via direct Course relationship on Timetable model
                compliance_sessions = AttendanceSession.objects.filter(timetable_entry__course__department_id=filter_dept)
                timetable_entries = TimetableEntry.objects.select_related('teacher__user').filter(course__department_id=filter_dept)
                
                if not timetable_entries.exists():
                    raise FieldError()
            except (FieldError, Exception):
                try:
                    # Tier 3 Lookup: Filter by the target Lecturer's home Department structural assignments
                    compliance_sessions = AttendanceSession.objects.filter(
                        Q(timetable_entry__teacher__department_id=filter_dept) | 
                        Q(timetable_entry__teacher__user__department_id=filter_dept)
                    )
                    timetable_entries = TimetableEntry.objects.select_related('teacher__user').filter(
                        Q(teacher__department_id=filter_dept) | 
                        Q(teacher__user__department_id=filter_dept)
                    )
                except (FieldError, Exception):
                    # Tier 4 Safe-Recovery Fallback: Maintain active structural roster list context
                    compliance_sessions = AttendanceSession.objects.all()
                    timetable_entries = TimetableEntry.objects.select_related('teacher__user').all()

    # Apply global secondary dashboard overrides if set
    if selected_stream_id and selected_stream_id != 'all':
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream_id=selected_stream_id)
        timetable_entries = timetable_entries.filter(stream_id=selected_stream_id)

    # Compute explicit metrics (ensures teachers show up even with 0 submission values)
    total_system_sessions = compliance_sessions.count()
    teacher_counts = compliance_sessions.values('timetable_entry__teacher_id').annotate(count=Count('id'))
    counts_map = {item['timetable_entry__teacher_id']: item['count'] for item in teacher_counts if item['timetable_entry__teacher_id']}

    teachers_raw = []
    seen_teachers = set()
    for entry in timetable_entries:
        if entry.teacher and entry.teacher.id not in seen_teachers:
            seen_teachers.add(entry.teacher.id)
            email = entry.teacher.user.email if entry.teacher.user else "—"
            submitted = counts_map.get(entry.teacher.id, 0)
            rate = round((submitted / total_system_sessions) * 100, 1) if total_system_sessions > 0 else 0
            teachers_raw.append({'email': email, 'submitted': submitted, 'rate': f"{rate}%"})

    top_submitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=True)[:20]
    top_unsubmitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=False)[:20]

    # Chart Generation Visual Processing Data Arrays
    chart_dist_data = [present_count, absent_count]
    streams_data = Stream.objects.annotate(
        total=Count('students__attendancerecord'),
        present=Count('students__attendancerecord', filter=Q(students__attendancerecord__status='PRESENT'))
    ).filter(total__gt=0)
    
    stream_labels = [stream.name for stream in streams_data]
    stream_rates = [round((stream.present / stream.total) * 100, 1) for stream in streams_data]

    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_present_students': top_present_students,
        'top_absent_students': top_absent_students,
        'top_submitted_teachers': top_submitted_teachers,
        'top_unsubmitted_teachers': top_unsubmitted_teachers,
        'streams': Stream.objects.all(),
        'selected_stream': selected_stream_id,
        'chart_dist_data': json.dumps(chart_dist_data),
        'stream_labels': json.dumps(stream_labels),
        'stream_rates': json.dumps(stream_rates),
        
        # Structural cascades UI framework mapping parameters
        'departments': all_departments,
        'courses': selectable_courses,
        'selectable_streams': selectable_streams,
        'detailed_student_reports': detailed_student_reports,
        'filter_dept': filter_dept,
        'filter_course': filter_course,
        'filter_stream': filter_stream,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'quick_range': quick_range,
    }
    return render(request, 'attendance/analytics_dashboard.html', context)
    # Enforce strict Admin-only access rule matrix
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # 1. CAPTURE DETAILED REPORT FILTERS
    filter_dept = request.GET.get('filter_dept')
    filter_course = request.GET.get('filter_course')
    filter_stream = request.GET.get('filter_stream')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    quick_range = request.GET.get('quick_range')

    # Base entity dropdown datasets supporting cascading display lists
    all_departments = Department.objects.all()
    selectable_courses = Course.objects.all()
    selectable_streams = Stream.objects.all()

    if filter_dept:
        selectable_courses = selectable_courses.filter(department_id=filter_dept)
    if filter_course:
        selectable_streams = selectable_streams.filter(course__code=filter_course)

    # 2. RESOLVE TIME DATA WINDOW CONSTRAINTS
    today = datetime.now().date()
    start_date = None
    end_date = None

    if quick_range == 'today':
        start_date = today
        end_date = today
    elif quick_range == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif quick_range == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    else:
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

    # Build cross-cutting performance filter arrays
    record_date_filter = Q()
    if start_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__gte=start_date)
    if end_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__lte=end_date)

    # 3. COMPILE COMPLETE DETAILED STUDENT POPULATION ROSTER
    report_students_queryset = StudentProfile.objects.select_related('stream', 'course__department').all()
    
    if filter_stream:
        report_students_queryset = report_students_queryset.filter(stream_id=filter_stream)
    elif filter_course:
        report_students_queryset = report_students_queryset.filter(course__code=filter_course)
    elif filter_dept:
        report_students_queryset = report_students_queryset.filter(course__department_id=filter_dept)

    # Calculate absolute percentages over selected date parameters
    annotated_students = report_students_queryset.annotate(
        total_filtered_records=Count('attendancerecord', filter=record_date_filter),
        present_filtered_records=Count('attendancerecord', filter=record_date_filter & Q(attendancerecord__status='PRESENT'))
    ).order_by('name')

    detailed_student_reports = []
    for s in annotated_students:
        tot = s.total_filtered_records
        pres = s.present_filtered_records
        percentage_val = round((pres / tot) * 100, 1) if tot > 0 else 0.0
        
        detailed_student_reports.append({
            "name": s.name,
            "reg_number": s.reg_number,
            "stream_name": s.stream.name if s.stream else "—",
            "total_logs": tot,
            "present_logs": pres,
            "attendance_percentage": f"{percentage_val}%",
            "percentage_num": percentage_val
        })

    # 4. INTERCEPT SPECIFIC DETAILED EXPORT REQUESTS
    if request.GET.get('export') == 'detailed_student_csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="detailed_student_report_{today}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Registration Number', 'Assigned Stream', 'Total Sessions Checked', 'Sessions Attended', 'Attendance Rate'])
        for item in detailed_student_reports:
            writer.writerow([item['name'], item['reg_number'], item['stream_name'], item['total_logs'], item['present_logs'], item['attendance_percentage']])
        return response

    # 5. PRESERVE ORIGINAL ANALYTICS COMPONENT ARRAYS
    selected_stream_id = request.GET.get('stream')
    records = AttendanceRecord.objects.all()
    sessions = AttendanceSession.objects.select_related('timetable_entry__teacher__user', 'timetable_entry__stream').order_by('-date_marked')
    
    if selected_stream_id and selected_stream_id != 'all':
        records = records.filter(student__stream_id=selected_stream_id)
        sessions = sessions.filter(timetable_entry__stream_id=selected_stream_id)

    # 6. INTERCEPT GLOBAL CSV EXPORT REQUEST
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Reg Number', 'Class/Stream', 'Status', 'Date'])
        for record in records.select_related('student__stream', 'session'):
            date_val = record.session.date_marked.strftime('%Y-%m-%d') if hasattr(record, 'session') and record.session.date_marked else 'N/A'
            stream_name = record.student.stream.name if record.student.stream else 'N/A'
            writer.writerow([record.student.name, record.student.reg_number, stream_name, record.status, date_val])
        return response

    # Process original global metrics
    total_records = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()

    stats = {
        'total_records': total_records,
        'present_rate': round((present_count / total_records) * 100, 1) if total_records > 0 else 0,
        'absent_rate': round((absent_count / total_records) * 100, 1) if total_records > 0 else 0,
    }

    # Location Log Parsing
    audit_logs = []
    for session in sessions[:20]:
        t_entry = session.timetable_entry
        has_coords = session.teacher_latitude is not None and session.teacher_longitude is not None
        location_str = f"{session.teacher_latitude}, {session.teacher_longitude}" if has_coords else "Not captured"
        audit_logs.append({
            "date": session.date_marked.strftime("%Y-%m-%d") if session.date_marked else "—",
            "class": t_entry.stream.name if t_entry.stream else "—",
            "lecturer": t_entry.teacher.user.email if t_entry.teacher and t_entry.teacher.user else "—",
            "location": location_str,
            "ip": "—",  
            "verified": has_coords  
        })

    # Standard Top-20 Matrix Ingestion 
    students = StudentProfile.objects.select_related('stream').annotate(
        total_rec=Count('attendancerecord'),
        present_rec=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT')),
        absent_rec=Count('attendancerecord', filter=Q(attendancerecord__status='ABSENT'))
    ).filter(total_rec__gt=0)
    
    students_raw = []
    for s in students:
        p_rate = (s.present_rec / s.total_rec) * 100
        a_rate = (s.absent_rec / s.total_rec) * 100
        students_raw.append({
            "name": s.name,
            "reg": s.reg_number,
            "class": s.stream.name if s.stream else "—",
            "p_rate_num": p_rate,
            "a_rate_num": a_rate,
            "p_rate": f"{round(p_rate, 1)}%",
            "a_rate": f"{round(a_rate, 1)}%"
        })
    
    top_present_students = sorted(students_raw, key=lambda x: x['p_rate_num'], reverse=True)[:20]
    top_absent_students = sorted(students_raw, key=lambda x: x['a_rate_num'], reverse=True)[:20]

    # 7. ADDED/UPDATED: Lecturer Attendance Submission Compliance Report Scope Filtering
    TimetableEntry = AttendanceSession._meta.get_field('timetable_entry').related_model
    
    compliance_sessions = AttendanceSession.objects.all()
    timetable_entries = TimetableEntry.objects.select_related('teacher__user').all()
    
    # Apply structural Department, Course, and Stream scope constraints dynamically
    if filter_stream:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream_id=filter_stream)
        timetable_entries = timetable_entries.filter(stream_id=filter_stream)
    elif filter_course:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream__course__code=filter_course)
        timetable_entries = timetable_entries.filter(stream__course__code=filter_course)
    elif filter_dept:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream__course__department_id=filter_dept)
        timetable_entries = timetable_entries.filter(stream__course__department_id=filter_dept)

    if selected_stream_id and selected_stream_id != 'all':
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream_id=selected_stream_id)
        timetable_entries = timetable_entries.filter(stream_id=selected_stream_id)

    total_system_sessions = compliance_sessions.count()
    teacher_counts = compliance_sessions.values('timetable_entry__teacher_id').annotate(count=Count('id'))
    counts_map = {item['timetable_entry__teacher_id']: item['count'] for item in teacher_counts if item['timetable_entry__teacher_id']}

    teachers_raw = []
    seen_teachers = set()
    for entry in timetable_entries:
        if entry.teacher and entry.teacher.id not in seen_teachers:
            seen_teachers.add(entry.teacher.id)
            email = entry.teacher.user.email if entry.teacher.user else "—"
            submitted = counts_map.get(entry.teacher.id, 0)
            rate = round((submitted / total_system_sessions) * 100, 1) if total_system_sessions > 0 else 0
            teachers_raw.append({'email': email, 'submitted': submitted, 'rate': f"{rate}%"})

    top_submitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=True)[:20]
    top_unsubmitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=False)[:20]

    # Chart Generation Data Processing
    chart_dist_data = [present_count, absent_count]
    streams_data = Stream.objects.annotate(
        total=Count('students__attendancerecord'),
        present=Count('students__attendancerecord', filter=Q(students__attendancerecord__status='PRESENT'))
    ).filter(total__gt=0)
    
    stream_labels = [stream.name for stream in streams_data]
    stream_rates = [round((stream.present / stream.total) * 100, 1) for stream in streams_data]

    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_present_students': top_present_students,
        'top_absent_students': top_absent_students,
        'top_submitted_teachers': top_submitted_teachers,
        'top_unsubmitted_teachers': top_unsubmitted_teachers,
        'streams': Stream.objects.all(),
        'selected_stream': selected_stream_id,
        'chart_dist_data': json.dumps(chart_dist_data),
        'stream_labels': json.dumps(stream_labels),
        'stream_rates': json.dumps(stream_rates),
        
        # Detailed reporting fields context
        'departments': all_departments,
        'courses': selectable_courses,
        'selectable_streams': selectable_streams,
        'detailed_student_reports': detailed_student_reports,
        'filter_dept': filter_dept,
        'filter_course': filter_course,
        'filter_stream': filter_stream,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'quick_range': quick_range,
    }
    return render(request, 'attendance/analytics_dashboard.html', context)
    # Enforce strict Admin-only access rule matrix
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # 1. CAPTURE DETAILED REPORT FILTERS
    filter_dept = request.GET.get('filter_dept')
    filter_course = request.GET.get('filter_course')
    filter_stream = request.GET.get('filter_stream')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    quick_range = request.GET.get('quick_range')

    # Base entity dropdown datasets supporting cascading display lists
    all_departments = Department.objects.all()
    selectable_courses = Course.objects.all()
    selectable_streams = Stream.objects.all()

    if filter_dept:
        selectable_courses = selectable_courses.filter(department_id=filter_dept)
    if filter_course:
        selectable_streams = selectable_streams.filter(course__code=filter_course)

    # 2. RESOLVE TIME DATA WINDOW CONSTRAINTS
    today = datetime.now().date()
    start_date = None
    end_date = None

    if quick_range == 'today':
        start_date = today
        end_date = today
    elif quick_range == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif quick_range == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    else:
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

    # Build cross-cutting performance filter arrays
    record_date_filter = Q()
    if start_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__gte=start_date)
    if end_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__lte=end_date)

    # 3. COMPILE COMPLETE DETAILED STUDENT POPULATION ROSTER
    report_students_queryset = StudentProfile.objects.select_related('stream', 'course__department').all()
    
    if filter_stream:
        report_students_queryset = report_students_queryset.filter(stream_id=filter_stream)
    elif filter_course:
        report_students_queryset = report_students_queryset.filter(course__code=filter_course)
    elif filter_dept:
        report_students_queryset = report_students_queryset.filter(course__department_id=filter_dept)

    # Calculate absolute percentages over selected date parameters
    annotated_students = report_students_queryset.annotate(
        total_filtered_records=Count('attendancerecord', filter=record_date_filter),
        present_filtered_records=Count('attendancerecord', filter=record_date_filter & Q(attendancerecord__status='PRESENT'))
    ).order_by('name')

    detailed_student_reports = []
    for s in annotated_students:
        tot = s.total_filtered_records
        pres = s.present_filtered_records
        percentage_val = round((pres / tot) * 100, 1) if tot > 0 else 0.0
        
        detailed_student_reports.append({
            "name": s.name,
            "reg_number": s.reg_number,
            "stream_name": s.stream.name if s.stream else "—",
            "total_logs": tot,
            "present_logs": pres,
            "attendance_percentage": f"{percentage_val}%",
            "percentage_num": percentage_val
        })

    # 4. INTERCEPT SPECIFIC DETAILED EXPORT REQUESTS
    if request.GET.get('export') == 'detailed_student_csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="detailed_student_report_{today}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Registration Number', 'Assigned Stream', 'Total Sessions Checked', 'Sessions Attended', 'Attendance Rate'])
        for item in detailed_student_reports:
            writer.writerow([item['name'], item['reg_number'], item['stream_name'], item['total_logs'], item['present_logs'], item['attendance_percentage']])
        return response

    # 5. PRESERVE ORIGINAL ANALYTICS COMPONENT ARRAYS
    selected_stream_id = request.GET.get('stream')
    records = AttendanceRecord.objects.all()
    sessions = AttendanceSession.objects.select_related('timetable_entry__teacher__user', 'timetable_entry__stream').order_by('-date_marked')
    
    if selected_stream_id and selected_stream_id != 'all':
        records = records.filter(student__stream_id=selected_stream_id)
        sessions = sessions.filter(timetable_entry__stream_id=selected_stream_id)

    # Process original global metrics
    total_records = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()

    stats = {
        'total_records': total_records,
        'present_rate': round((present_count / total_records) * 100, 1) if total_records > 0 else 0,
        'absent_rate': round((absent_count / total_records) * 100, 1) if total_records > 0 else 0,
    }

    # Location Log Parsing
    audit_logs = []
    for session in sessions[:20]:
        t_entry = session.timetable_entry
        has_coords = session.teacher_latitude is not None and session.teacher_longitude is not None
        location_str = f"{session.teacher_latitude}, {session.teacher_longitude}" if has_coords else "Not captured"
        audit_logs.append({
            "date": session.date_marked.strftime("%Y-%m-%d") if session.date_marked else "—",
            "class": t_entry.stream.name if t_entry.stream else "—",
            "lecturer": t_entry.teacher.user.email if t_entry.teacher and t_entry.teacher.user else "—",
            "location": location_str,
            "ip": "—",  
            "verified": has_coords  
        })

    # Standard Top-20 Matrix Ingestion 
    students = StudentProfile.objects.select_related('stream').annotate(
        total_rec=Count('attendancerecord'),
        present_rec=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT')),
        absent_rec=Count('attendancerecord', filter=Q(attendancerecord__status='ABSENT'))
    ).filter(total_rec__gt=0)
    
    students_raw = []
    for s in students:
        p_rate = (s.present_rec / s.total_rec) * 100
        a_rate = (s.absent_rec / s.total_rec) * 100
        students_raw.append({
            "name": s.name,
            "reg": s.reg_number,
            "class": s.stream.name if s.stream else "—",
            "p_rate_num": p_rate,
            "a_rate_num": a_rate,
            "p_rate": f"{round(p_rate, 1)}%",
            "a_rate": f"{round(a_rate, 1)}%"
        })
    
    top_present_students = sorted(students_raw, key=lambda x: x['p_rate_num'], reverse=True)[:20]
    top_absent_students = sorted(students_raw, key=lambda x: x['a_rate_num'], reverse=True)[:20]

    # Lecturer Compliance Ingestion
    TimetableEntry = AttendanceSession._meta.get_field('timetable_entry').related_model
    total_system_sessions = sessions.count()
    teacher_counts = sessions.values('timetable_entry__teacher_id').annotate(count=Count('id'))
    counts_map = {item['timetable_entry__teacher_id']: item['count'] for item in teacher_counts if item['timetable_entry__teacher_id']}
    timetable_entries = TimetableEntry.objects.select_related('teacher__user').all()
    
    if selected_stream_id and selected_stream_id != 'all':
        timetable_entries = timetable_entries.filter(stream_id=selected_stream_id)

    teachers_raw = []
    seen_teachers = set()
    for entry in timetable_entries:
        if entry.teacher and entry.teacher.id not in seen_teachers:
            seen_teachers.add(entry.teacher.id)
            email = entry.teacher.user.email if entry.teacher.user else "—"
            submitted = counts_map.get(entry.teacher.id, 0)
            rate = round((submitted / total_system_sessions) * 100, 1) if total_system_sessions > 0 else 0
            teachers_raw.append({'email': email, 'submitted': submitted, 'rate': f"{rate}%"})

    top_submitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=True)[:20]
    top_unsubmitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=False)[:20]

    # Chart Generation Data Processing
    chart_dist_data = [present_count, absent_count]
    streams_data = Stream.objects.annotate(
        total=Count('students__attendancerecord'),
        present=Count('students__attendancerecord', filter=Q(students__attendancerecord__status='PRESENT'))
    ).filter(total__gt=0)
    
    stream_labels = [stream.name for stream in streams_data]
    stream_rates = [round((stream.present / stream.total) * 100, 1) for stream in streams_data]

    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_present_students': top_present_students,
        'top_absent_students': top_absent_students,
        'top_submitted_teachers': top_submitted_teachers,
        'top_unsubmitted_teachers': top_unsubmitted_teachers,
        'streams': Stream.objects.all(),
        'selected_stream': selected_stream_id,
        'chart_dist_data': json.dumps(chart_dist_data),
        'stream_labels': json.dumps(stream_labels),
        'stream_rates': json.dumps(stream_rates),
        
        # New Detailed reporting fields injected safely into context context
        'departments': all_departments,
        'courses': selectable_courses,
        'selectable_streams': selectable_streams,
        'detailed_student_reports': detailed_student_reports,
        'filter_dept': filter_dept,
        'filter_course': filter_course,
        'filter_stream': filter_stream,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'quick_range': quick_range,
    }
    return render(request, 'attendance/analytics_dashboard.html', context)
    # Enforce strict Admin-only access rule matrix
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # Extract optional filter parameters from template request scopes
    selected_stream_id = request.GET.get('stream')
    
    # 1. INITIALIZE VARIABLES FIRST (Fixes UnboundLocalError)
    # Base query sets targeting core transactional datasets
    records = AttendanceRecord.objects.all()
    sessions = AttendanceSession.objects.select_related(
        'timetable_entry__teacher__user', 
        'timetable_entry__stream'
    ).order_by('-date_marked')
    
    # 2. APPLY SCOPING FILTERS dynamically if requested
    if selected_stream_id and selected_stream_id != 'all':
        records = records.filter(student__stream_id=selected_stream_id)
        sessions = sessions.filter(timetable_entry__stream_id=selected_stream_id)

    # 3. INTERCEPT CSV EXPORT REQUEST
    # Now that 'records' exists and is filtered, we can safely export it!
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Reg Number', 'Class/Stream', 'Status', 'Date'])
        
        # NOTE: If your relation to AttendanceSession is named something else 
        # (e.g. 'attendance_session'), change 'session' below to match your models.py
        for record in records.select_related('student__stream', 'session'):
            date_val = record.session.date_marked.strftime('%Y-%m-%d') if hasattr(record, 'session') and record.session.date_marked else 'N/A'
            stream_name = record.student.stream.name if record.student.stream else 'N/A'
            
            writer.writerow([
                record.student.name, 
                record.student.reg_number,
                stream_name,
                record.status, 
                date_val
            ])
        return response

    # 4. Compute Live Summary Statistics Metrics Matrix
    total_records = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()

    stats = {
        'total_records': total_records,
        'present_rate': round((present_count / total_records) * 100, 1) if total_records > 0 else 0,
        'absent_rate': round((absent_count / total_records) * 100, 1) if total_records > 0 else 0,
    }

    # 5. Compile Lecturer Location Audit Log Dataset Dynamically
    audit_logs = []
    for session in sessions[:20]:
        t_entry = session.timetable_entry
        has_coords = session.teacher_latitude is not None and session.teacher_longitude is not None
        location_str = f"{session.teacher_latitude}, {session.teacher_longitude}" if has_coords else "Not captured"
        
        audit_logs.append({
            "date": session.date_marked.strftime("%Y-%m-%d") if session.date_marked else "—",
            "class": t_entry.stream.name if t_entry.stream else "—",
            "lecturer": t_entry.teacher.user.email if t_entry.teacher and t_entry.teacher.user else "—",
            "location": location_str,
            "ip": "—",  
            "verified": has_coords  
        })

    # 6. Pull Top Present & Absent Students Metrics Arrays via Database Annotations (Top 20)
    students = StudentProfile.objects.select_related('stream').annotate(
        total_rec=Count('attendancerecord'),
        present_rec=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT')),
        absent_rec=Count('attendancerecord', filter=Q(attendancerecord__status='ABSENT'))
    ).filter(total_rec__gt=0)
    
    students_raw = []
    for s in students:
        p_rate = (s.present_rec / s.total_rec) * 100
        a_rate = (s.absent_rec / s.total_rec) * 100
        students_raw.append({
            "name": s.name,
            "reg": s.reg_number,
            "class": s.stream.name if s.stream else "—",
            "p_rate_num": p_rate,
            "a_rate_num": a_rate,
            "p_rate": f"{round(p_rate, 1)}%",
            "a_rate": f"{round(a_rate, 1)}%"
        })
    
    top_present_students = sorted(students_raw, key=lambda x: x['p_rate_num'], reverse=True)[:20]
    top_absent_students = sorted(students_raw, key=lambda x: x['a_rate_num'], reverse=True)[:20]

    # 7. Generate Teacher Attendance Submission Compliance Metrics Report
    TimetableEntry = AttendanceSession._meta.get_field('timetable_entry').related_model
    total_system_sessions = sessions.count()

    teacher_counts = sessions.values('timetable_entry__teacher_id').annotate(count=Count('id'))
    counts_map = {item['timetable_entry__teacher_id']: item['count'] for item in teacher_counts if item['timetable_entry__teacher_id']}

    timetable_entries = TimetableEntry.objects.select_related('teacher__user').all()
    if selected_stream_id and selected_stream_id != 'all':
        timetable_entries = timetable_entries.filter(stream_id=selected_stream_id)

    teachers_raw = []
    seen_teachers = set()

    for entry in timetable_entries:
        if entry.teacher and entry.teacher.id not in seen_teachers:
            seen_teachers.add(entry.teacher.id)
            email = entry.teacher.user.email if entry.teacher.user else "—"
            submitted = counts_map.get(entry.teacher.id, 0)
            
            rate = round((submitted / total_system_sessions) * 100, 1) if total_system_sessions > 0 else 0
            
            teachers_raw.append({
                'email': email,
                'submitted': submitted,
                'rate': f"{rate}%"
            })

    top_submitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=True)[:20]
    top_unsubmitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=False)[:20]

    # 8. Generate Aggregated Payload Engines for ChartJS
    chart_dist_data = [present_count, absent_count]
    
    streams_data = Stream.objects.annotate(
        total=Count('students__attendancerecord'),
        present=Count('students__attendancerecord', filter=Q(students__attendancerecord__status='PRESENT'))
    ).filter(total__gt=0)
    
    stream_labels = [stream.name for stream in streams_data]
    stream_rates = [round((stream.present / stream.total) * 100, 1) for stream in streams_data]

    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_present_students': top_present_students,
        'top_absent_students': top_absent_students,
        'top_submitted_teachers': top_submitted_teachers,
        'top_unsubmitted_teachers': top_unsubmitted_teachers,
        'streams': Stream.objects.all(),
        'selected_stream': selected_stream_id,
        'chart_dist_data': json.dumps(chart_dist_data),
        'stream_labels': json.dumps(stream_labels),
        'stream_rates': json.dumps(stream_rates),
    }
    return render(request, 'attendance/analytics_dashboard.html', context)



from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from .models import TimetableBatch, TimetableEntry

@login_required
def export_timetable_pdf(request):
    """
    Generates a formal landscape PDF matrix tracking the current 
    active master schedule layout configurations.
    """
    # 1. Fetch the exact same active master batch as the editor view
    batch = TimetableBatch.objects.filter(is_active=True, is_revoked=False).order_by('-uploaded_at').first() #
    if not batch:
        messages.warning(request, "No active timetable configuration was found to generate a document.")
        return redirect('attendance:upload_timetable')
        
    # 2. Initialize HTTP Response object configured as PDF payload stream
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="timetable_batch_{batch.id}.pdf"' #
    
    # Setup document geometry in landscape format to accommodate the 7 day tracks
    doc = SimpleDocTemplate(
        response, 
        pagesize=landscape(letter), 
        rightMargin=30, 
        leftMargin=30, 
        topMargin=35, 
        bottomMargin=30
    )
    story = []
    
    # 3. Setup Layout Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#2b3e85'),
        alignment=1, # Centered alignment
        spaceAfter=15
    )
    
    cell_text_style = ParagraphStyle(
        'MatrixCell',
        parent=styles['Normal'],
        fontSize=7.5,
        leading=10,
        alignment=1 # Center text wrapped within table nodes
    )
    
    header_style = ParagraphStyle(
        'MatrixHeader',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        fontName='Helvetica-Bold',
        textColor=colors.whitesmoke,
        alignment=1
    )
    
    # Document header markup text block
    formatted_date = batch.week_start_date.strftime('%B %d, %Y') #
    story.append(Paragraph(f"MASTER TIMETABLE SCHEDULE GRID", title_style))
    story.append(Paragraph(f"Active Schedule Target Cycle — Week Commencing: {formatted_date}", ParagraphStyle('Sub', alignment=1, fontSize=10, textColor=colors.HexColor('#475569'))))
    story.append(Spacer(1, 20))
    
    # 4. Extract and Group DB Matrix Entries (identical workflow as editor GET request)
    DAYS = [code for code, _ in TimetableEntry.DAYS_OF_WEEK] #
    headers = [Paragraph("Time Block Interval", header_style)] + [Paragraph(name, header_style) for _, name in TimetableEntry.DAYS_OF_WEEK] #
    
    entries = TimetableEntry.objects.filter(batch=batch).order_by('start_time') #
    grouped_slots = {}
    for entry in entries: #
        time_key = (entry.start_time.strftime('%H:%M'), entry.end_time.strftime('%H:%M')) #
        if time_key not in grouped_slots:
            grouped_slots[time_key] = {}
        grouped_slots[time_key][entry.day] = entry #
        
    matrix_data = [headers]
    
    # Build structural rows tracking each unique coordinate matrix node cell block
    for (start, end), day_map in grouped_slots.items():
        row_cells = [Paragraph(f"<b>{start} - {end}</b>", cell_text_style)]
        
        for day_code in DAYS:
            entry = day_map.get(day_code)
            if entry:
                # Text structural wrapping to safely output multi-line labels inside table cell nodes
                cell_content = f"<b>{entry.course_unit.code}</b><br/>{entry.teacher.name}<br/><font color='#475569'>{entry.stream.name}</font>" #
                row_cells.append(Paragraph(cell_content, cell_text_style))
            else:
                row_cells.append(Paragraph("<font color='#cbd5e1'>—</font>", cell_text_style))
                
        matrix_data.append(row_cells)
        
    # Calculate geometric column constraints to match page printable canvas space boundaries
    # Printable horizontal target width: 792 (landscape width) - 60 (margins) = 732 points total
    col_widths = [90] + [91] * 7 
    
    timetable_table = Table(matrix_data, colWidths=col_widths, repeatRows=1)
    timetable_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b3e85')), # Match original navy headers
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')), # Subtle structural slate boundaries
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    story.append(timetable_table)
    
    # 5. Compile payload contents out to the client download pipeline stream
    doc.build(story)
    return response


@login_required
@transaction.atomic
def manage_users(request):
    # Guard to ensure only ADMIN can view or change roles
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
        
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        new_role = request.POST.get('role')
        
        if user_id and new_role in dict(User.ROLE_CHOICES):
            target_user = get_object_or_404(User, id=user_id)
            target_user.role = new_role
            target_user.save()
            messages.success(request, f"Updated role for {target_user.username} to {new_role}.")
            return redirect('attendance:manage_users')

    all_users = User.objects.all().order_by('username')
    return render(request, 'attendance/manage_users.html', {
        'all_users': all_users,
        'role_choices': User.ROLE_CHOICES
    })














#----------------------------------------------------------------------------------
# ==================== EXAMINATION MODULE (Admin/Registrar) ====================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.db.models import Avg, Count, Q, F
from .models import (
    User, Exam, GradeScale, MarksEntry, Transcript, StudentProfile,
    CourseUnit, AcademicTerm, Stream
)
from .forms import ExamForm, GradeScaleForm, MarksEntryForm  # define these forms as needed
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.utils import timezone

from attendance.models import User, AcademicTerm, CourseUnit, Exam, Course


from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.utils import timezone

from .models import User, AcademicTerm, CourseUnit, Exam, Course


@login_required
@transaction.atomic
def manage_exams(request):
    """
    Manages exam creation and displays the exam timetable schedule,
    hierarchically organized by Course (Program) -> Course Unit -> Exams.
    Supports filtering by selected term or defaulting to active term.
    """
    if request.user.role not in [User.IS_ADMIN, User.IS_REGISTRAR, User.IS_TEACHER]:
        return HttpResponse("Unauthorized", status=403)

    # 1. Fetch active term and terms list
    active_term = AcademicTerm.objects.filter(is_current=True).first()
    all_terms = AcademicTerm.objects.all().order_by('-start_date')

    # Determine which term to display (GET param or default to active term)
    selected_term_id = request.GET.get('term')
    if selected_term_id:
        selected_term = AcademicTerm.objects.filter(id=selected_term_id).first()
    else:
        selected_term = active_term

    # 2. Handle POST Request to Create a New Exam
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        course_unit_code = request.POST.get('course_unit', '').strip()
        exam_date = request.POST.get('exam_date', '').strip()
        start_time = request.POST.get('start_time', '').strip()
        end_time = request.POST.get('end_time', '').strip()
        total_marks = request.POST.get('total_marks', 100)
        is_published = request.POST.get('is_published') == 'on'

        target_term = selected_term or active_term

        if not target_term:
            messages.error(request, "Failed to schedule exam: No valid Academic Term is selected or active.")
            return redirect('attendance:manage_exams')

        if name and course_unit_code and exam_date and start_time and end_time:
            try:
                course_unit = CourseUnit.objects.get(code=course_unit_code)
                
                Exam.objects.create(
                    name=name,
                    course_unit=course_unit,
                    term=target_term,
                    exam_date=exam_date,
                    start_time=start_time,
                    end_time=end_time,
                    total_marks=total_marks,
                    is_published=is_published
                )
                messages.success(request, f"Exam '{name}' scheduled successfully for '{target_term}'.")
            except CourseUnit.DoesNotExist:
                messages.error(request, "The selected course unit does not exist.")
            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
        else:
            messages.error(request, "Please fill in all required timetable details.")

        return redirect(f"{request.path}?term={target_term.id}" if target_term else request.path)

    # 3. Retrieve Exams for Selected Term
    exams = Exam.objects.select_related(
        'course_unit',
        'course_unit__course',
        'term'
    )
    if selected_term:
        exams = exams.filter(term=selected_term)

    exams = exams.order_by('course_unit__course__name', 'course_unit__code', 'exam_date', 'start_time')

    now = timezone.now()
    current_date = now.date()
    current_time = now.time()

    # 4. Group Exams: Course -> Course Unit -> Exams with status
    grouped_data = {}
    for exam in exams:
        if exam.exam_date < current_date or (exam.exam_date == current_date and current_time > exam.end_time):
            status = 'DONE'
        elif exam.exam_date == current_date and exam.start_time <= current_time <= exam.end_time:
            status = 'IN_PROGRESS'
        else:
            status = 'UPCOMING'

        course = exam.course_unit.course
        unit = exam.course_unit

        if course not in grouped_data:
            grouped_data[course] = {}
        if unit not in grouped_data[course]:
            grouped_data[course][unit] = []

        grouped_data[course][unit].append({
            'exam': exam,
            'status': status
        })

    # Format into a clean structured list for Django template iteration
    courses_with_exams = []
    for course, units_dict in grouped_data.items():
        units_list = []
        for unit, exam_list in units_dict.items():
            units_list.append({
                'unit': unit,
                'exams': exam_list
            })
        courses_with_exams.append({
            'course': course,
            'units': units_list
        })

    course_units = CourseUnit.objects.select_related('course').all()
    courses = Course.objects.all()

    return render(request, 'attendance/manage_exams.html', {
        'courses_with_exams': courses_with_exams,
        'active_term': active_term,
        'selected_term': selected_term,
        'terms': all_terms,
        'courses': courses,
        'course_units': course_units,
    })

@login_required
@transaction.atomic
def edit_exam(request, exam_id):  # Fixed parameter name from pk to exam_id
    """
    Edits existing exam entries while maintaining active term assignments and updated timetable slots.
    """
    if request.user.role not in [User.IS_ADMIN, User.IS_REGISTRAR, User.IS_TEACHER]:
        return HttpResponse("Unauthorized", status=403)

    exam = get_object_or_404(Exam, pk=exam_id)
    active_term = AcademicTerm.objects.filter(is_current=True).first()

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        course_unit_code = request.POST.get('course_unit', '').strip()
        exam_date = request.POST.get('exam_date', '').strip()
        start_time = request.POST.get('start_time', '').strip()
        end_time = request.POST.get('end_time', '').strip()
        total_marks = request.POST.get('total_marks', 100)
        is_published = request.POST.get('is_published') == 'on'

        if name and course_unit_code and exam_date and start_time and end_time:
            try:
                course_unit = CourseUnit.objects.get(code=course_unit_code)
                
                exam.name = name
                exam.course_unit = course_unit
                if active_term:
                    exam.term = active_term
                exam.exam_date = exam_date
                exam.start_time = start_time
                exam.end_time = end_time
                exam.total_marks = total_marks
                exam.is_published = is_published
                exam.save()

                messages.success(request, f"Timetable entry for exam '{exam.name}' updated successfully.")
                return redirect('attendance:manage_exams')
            except CourseUnit.DoesNotExist:
                messages.error(request, "Selected course unit does not exist.")
            except Exception as e:
                messages.error(request, f"Error updating exam timetable: {str(e)}")
        else:
            messages.error(request, "Please ensure all required fields are filled.")

    course_units = CourseUnit.objects.select_related('course').all()
    return render(request, 'attendance/edit_exam.html', {
        'exam': exam,
        'active_term': active_term,
        'course_units': course_units,
    })


@login_required
@transaction.atomic
def delete_exam(request, exam_id):  # Fixed parameter name from pk to exam_id
    """
    Deletes an exam entry from the timetable.
    """
    if request.user.role not in [User.IS_ADMIN, User.IS_REGISTRAR, User.IS_TEACHER]:
        return HttpResponse("Unauthorized", status=403)

    exam = get_object_or_404(Exam, pk=exam_id)
    exam_name = exam.name
    exam.delete()
    
    messages.success(request, f"Exam '{exam_name}' removed from the exam timetable.")
    return redirect('attendance:manage_exams')

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import User, GradeScale, MarksEntry
from .forms import GradeScaleForm  # or wherever GradeScaleForm is defined

@login_required
def view_grades(request):
    """Display grade scales and auto-grade marks (Admin/Registrar)."""
    if request.user.role not in [User.IS_ADMIN, User.IS_REGISTRAR]:
        return HttpResponse("Unauthorized", status=403)

    scales = GradeScale.objects.all()

    if request.method == 'POST':
        action = request.POST.get('action')

        # Handle deletion
        if action == 'delete':
            scale_id = request.POST.get('scale_id')
            scale = get_object_or_404(GradeScale, id=scale_id)
            scale_name = scale.name
            scale.delete()
            messages.success(request, f"Grade scale '{scale_name}' deleted successfully.")
            return redirect('attendance:view_grades')

        # Handle grade scale creation/update
        scale_id = request.POST.get('scale_id')
        if scale_id:
            scale = get_object_or_404(GradeScale, id=scale_id)
            form = GradeScaleForm(request.POST, instance=scale)
        else:
            form = GradeScaleForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Grade scale saved successfully.")
            return redirect('attendance:view_grades')
    else:
        form = GradeScaleForm()

    # Auto-grade existing marks if needed
    if request.GET.get('auto_grade'):
        marks = MarksEntry.objects.filter(grade__isnull=True)
        for mark in marks:
            grade = GradeScale.objects.filter(
                min_score__lte=mark.marks_obtained,
                max_score__gte=mark.marks_obtained
            ).first()
            if grade:
                mark.grade = grade
                mark.save()
        messages.success(request, f"Auto-graded {marks.count()} marks.")
        return redirect('attendance:view_grades')

    return render(request, 'attendance/view_grades.html', {
        'scales': scales,
        'form': form,
    })

@login_required
@transaction.atomic
def generate_transcript(request, student_id, term_id):
    """Generate official transcript PDF (Registrar/Admin)."""
    if request.user.role not in [User.IS_ADMIN, User.IS_REGISTRAR]:
        return HttpResponse("Unauthorized", status=403)

    student = get_object_or_404(StudentProfile, reg_number=student_id)
    term = get_object_or_404(AcademicTerm, id=term_id)

    # Check if transcript already exists
    transcript, created = Transcript.objects.get_or_create(student=student, term=term)

    if created:
        # Generate PDF content - you can use reportlab or weasyprint
        # For simplicity, we'll create a placeholder PDF using reportlab
        from reportlab.pdfgen import canvas
        from io import BytesIO
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=(595.27, 841.89))  # A4
        p.drawString(100, 800, f"TRANSCRIPT FOR {student.name}")
        p.drawString(100, 780, f"Registration: {student.reg_number}")
        p.drawString(100, 760, f"Course: {student.course.code} - {student.course.name}")
        p.drawString(100, 740, f"Term: {term.academic_year} - {term.get_term_display()}")
        # ... add more content (marks, grades)
        p.showPage()
        p.save()
        buffer.seek(0)
        # Save PDF to transcript.pdf_file field
        from django.core.files.base import ContentFile
        transcript.pdf_file.save(f"transcript_{student.reg_number}_{term.id}.pdf", ContentFile(buffer.read()))
        buffer.close()
        messages.success(request, "Transcript generated successfully.")

    # Redirect to download or view
    if transcript.pdf_file:
        return redirect(transcript.pdf_file.url)
    else:
        messages.error(request, "Transcript PDF not available.")
        return redirect('attendance:registrar_dashboard')


from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Avg, Sum
from .models import User, Exam, Stream, CourseUnit, StudentProfile

@login_required
def exam_ranking(request):
    """
    View class/course rankings (Teacher/Admin/Registrar).
    Defaults to overall ranking across all exams, ordered by top performance 
    (Number of A's, B's, Average Marks, Total Marks).
    """
    if request.user.role not in [User.IS_ADMIN, User.IS_TEACHER, User.IS_REGISTRAR]:
        return HttpResponse("Unauthorized", status=403)

    # Get filter parameters
    exam_id = request.GET.get('exam', '').strip()
    stream_id = request.GET.get('stream', '').strip()
    course_unit_id = request.GET.get('course_unit', '').strip()

    exams = Exam.objects.select_related('course_unit').all()
    streams = Stream.objects.all()
    course_units = CourseUnit.objects.all()

    # Query students who have exam marks
    students_qs = StudentProfile.objects.filter(marks__isnull=False)

    # Apply optional filter conditions to the marks queryset
    if stream_id:
        students_qs = students_qs.filter(stream_id=stream_id)
    if course_unit_id:
        students_qs = students_qs.filter(marks__exam__course_unit_id=course_unit_id)
    if exam_id:
        students_qs = students_qs.filter(marks__exam_id=exam_id)

    # Perform conditional aggregation to calculate grade counts and score statistics
    students_qs = students_qs.annotate(
        total_exams=Count('marks', distinct=True),
        total_marks=Sum('marks__marks_obtained'),
        avg_marks=Avg('marks__marks_obtained'),
        count_a=Count('marks', filter=Q(marks__grade__name__iexact='A')),
        count_b=Count('marks', filter=Q(marks__grade__name__iexact='B')),
        count_c=Count('marks', filter=Q(marks__grade__name__iexact='C')),
        count_d=Count('marks', filter=Q(marks__grade__name__iexact='D')),
        count_f=Count('marks', filter=Q(marks__grade__name__iexact='F')),
    ).distinct()

    # Default performance sorting: Top A's -> Top B's -> Higher Average -> Higher Total Marks
    students_qs = students_qs.order_by('-count_a', '-count_b', '-avg_marks', '-total_marks')

    rankings = []
    for idx, student in enumerate(students_qs, start=1):
        rankings.append({
            'rank': idx,
            'student': student.name,
            'reg_number': student.reg_number,
            'stream': student.stream.name if student.stream else '-',
            'total_exams': student.total_exams,
            'avg_marks': round(student.avg_marks or 0, 1),
            'total_marks': round(student.total_marks or 0, 1),
            'count_a': student.count_a,
            'count_b': student.count_b,
            'count_c': student.count_c,
            'count_d': student.count_d,
            'count_f': student.count_f,
        })

    context = {
        'exams': exams,
        'streams': streams,
        'course_units': course_units,
        'rankings': rankings,
        'selected_exam': exam_id,
        'selected_stream': stream_id,
        'selected_course_unit': course_unit_id,
    }
    return render(request, 'attendance/exam_ranking.html', context)
# ==================== INVENTORY MODULE ====================

from .models import InventoryItem, Supplier, Asset, Procurement, StockMovement

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import User, InventoryItem, Supplier, Asset

@login_required
def manage_inventory(request):
    """CRUD for inventory items, assets, suppliers."""
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    items = InventoryItem.objects.select_related('supplier').all()
    suppliers = Supplier.objects.all()
    assets = Asset.objects.select_related('item', 'assigned_to').all()
    users = User.objects.filter(is_active=True)

    if request.method == 'POST':
        action = request.POST.get('action')

        # --- INVENTORY ITEM ACTIONS ---
        if action == 'add_item':
            supplier_id = request.POST.get('supplier')
            supplier = Supplier.objects.filter(id=supplier_id).first() if supplier_id else None
            InventoryItem.objects.create(
                name=request.POST.get('name'),
                category=request.POST.get('category'),
                quantity=request.POST.get('quantity') or 0,
                unit_price=request.POST.get('unit_price') or None,
                supplier=supplier,
                location=request.POST.get('location'),
                reorder_level=request.POST.get('reorder_level') or None
            )
            messages.success(request, "Inventory item created successfully.")

        elif action == 'edit_item':
            item_id = request.POST.get('item_id')
            item = get_object_or_404(InventoryItem, id=item_id)
            supplier_id = request.POST.get('supplier')
            item.name = request.POST.get('name')
            item.category = request.POST.get('category')
            item.quantity = request.POST.get('quantity') or 0
            item.unit_price = request.POST.get('unit_price') or None
            item.supplier = Supplier.objects.filter(id=supplier_id).first() if supplier_id else None
            item.location = request.POST.get('location')
            item.reorder_level = request.POST.get('reorder_level') or None
            item.save()
            messages.success(request, f"Item '{item.name}' updated successfully.")

        elif action == 'delete_item':
            item_id = request.POST.get('item_id')
            item = get_object_or_404(InventoryItem, id=item_id)
            item_name = item.name
            item.delete()
            messages.success(request, f"Item '{item_name}' deleted successfully.")

        # --- SUPPLIER ACTIONS ---
        elif action == 'add_supplier':
            Supplier.objects.create(
                name=request.POST.get('name'),
                contact_person=request.POST.get('contact_person'),
                phone=request.POST.get('phone'),
                email=request.POST.get('email'),
                address=request.POST.get('address')
            )
            messages.success(request, "Supplier added successfully.")

        elif action == 'edit_supplier':
            sup_id = request.POST.get('supplier_id')
            supplier = get_object_or_404(Supplier, id=sup_id)
            supplier.name = request.POST.get('name')
            supplier.contact_person = request.POST.get('contact_person')
            supplier.phone = request.POST.get('phone')
            supplier.email = request.POST.get('email')
            supplier.address = request.POST.get('address')
            supplier.save()
            messages.success(request, f"Supplier '{supplier.name}' updated successfully.")

        elif action == 'delete_supplier':
            sup_id = request.POST.get('supplier_id')
            supplier = get_object_or_404(Supplier, id=sup_id)
            sup_name = supplier.name
            supplier.delete()
            messages.success(request, f"Supplier '{sup_name}' deleted successfully.")

        # --- ASSET ACTIONS ---
        elif action == 'add_asset':
            item_id = request.POST.get('item')
            assigned_to_id = request.POST.get('assigned_to')
            item = get_object_or_404(InventoryItem, id=item_id)
            assigned_to = User.objects.filter(id=assigned_to_id).first() if assigned_to_id else None
            Asset.objects.create(
                asset_tag=request.POST.get('asset_tag'),
                item=item,
                serial_number=request.POST.get('serial_number'),
                assigned_to=assigned_to,
                status=request.POST.get('status', 'AVAILABLE')
            )
            messages.success(request, "Asset registered successfully.")

        elif action == 'edit_asset':
            asset_id = request.POST.get('asset_id')
            asset = get_object_or_404(Asset, id=asset_id)
            item_id = request.POST.get('item')
            assigned_to_id = request.POST.get('assigned_to')
            asset.item = get_object_or_404(InventoryItem, id=item_id)
            asset.assigned_to = User.objects.filter(id=assigned_to_id).first() if assigned_to_id else None
            asset.asset_tag = request.POST.get('asset_tag')
            asset.serial_number = request.POST.get('serial_number')
            asset.status = request.POST.get('status', 'AVAILABLE')
            asset.save()
            messages.success(request, f"Asset '{asset.asset_tag}' updated successfully.")

        elif action == 'delete_asset':
            asset_id = request.POST.get('asset_id')
            asset = get_object_or_404(Asset, id=asset_id)
            tag = asset.asset_tag
            asset.delete()
            messages.success(request, f"Asset '{tag}' deleted successfully.")

        return redirect('attendance:manage_inventory')

    return render(request, 'attendance/manage_inventory.html', {
        'items': items,
        'suppliers': suppliers,
        'assets': assets,
        'users': users,
    })

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse

@login_required
def manage_procurement(request):
    """Create, approve, edit, and delete procurement requests."""
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    procurements = Procurement.objects.select_related('item', 'supplier', 'approved_by').all()
    items = InventoryItem.objects.all()
    suppliers = Supplier.objects.all()

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            item_id = request.POST.get('item')
            quantity = request.POST.get('quantity')
            unit_cost = request.POST.get('unit_cost')
            supplier_id = request.POST.get('supplier')
            expected_delivery = request.POST.get('expected_delivery') or None
            
            item = get_object_or_404(InventoryItem, id=item_id)
            supplier = Supplier.objects.get(id=supplier_id) if supplier_id else None
            
            Procurement.objects.create(
                item=item,
                quantity=quantity,
                unit_cost=unit_cost,
                supplier=supplier,
                expected_delivery=expected_delivery,
                status='REQUESTED'
            )
            messages.success(request, "Procurement request created successfully.")

        elif action == 'approve':
            proc_id = request.POST.get('proc_id')
            proc = get_object_or_404(Procurement, id=proc_id)
            proc.status = 'APPROVED'
            proc.approved_by = request.user
            proc.save()
            messages.success(request, "Procurement request approved.")

        elif action == 'edit':
            proc_id = request.POST.get('proc_id')
            proc = get_object_or_404(Procurement, id=proc_id)
            
            item_id = request.POST.get('item')
            supplier_id = request.POST.get('supplier')
            
            proc.item = get_object_or_404(InventoryItem, id=item_id)
            proc.quantity = request.POST.get('quantity')
            proc.unit_cost = request.POST.get('unit_cost')
            proc.supplier = Supplier.objects.get(id=supplier_id) if supplier_id else None
            
            expected_delivery = request.POST.get('expected_delivery')
            if expected_delivery:
                proc.expected_delivery = expected_delivery
                
            status = request.POST.get('status')
            if status:
                proc.status = status
                
            proc.save()
            messages.success(request, "Procurement request updated successfully.")

        elif action == 'delete':
            proc_id = request.POST.get('proc_id')
            proc = get_object_or_404(Procurement, id=proc_id)
            proc.delete()
            messages.success(request, "Procurement request deleted successfully.")

        return redirect('attendance:manage_procurement')

    return render(request, 'attendance/manage_procurement.html', {
        'procurements': procurements,
        'items': items,
        'suppliers': suppliers,
    })

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse

@login_required
def asset_tracking(request):
    """Assign assets to staff, update status, or delete assets."""
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        action = request.POST.get('action', 'update')
        asset_id = request.POST.get('asset_id')
        asset = get_object_or_404(Asset, id=asset_id)

        if action == 'delete':
            asset_name = str(asset)
            asset.delete()
            messages.success(request, f"Asset '{asset_name}' deleted successfully.")
        else:
            user_id = request.POST.get('assigned_to')
            status = request.POST.get('status')
            
            asset.assigned_to = User.objects.get(id=user_id) if user_id else None
            asset.status = status
            asset.save()
            messages.success(request, "Asset updated successfully.")

        return redirect('attendance:asset_tracking')

    assets = Asset.objects.select_related('item', 'assigned_to').all()
    staff_users = User.objects.exclude(role=User.IS_STUDENT)

    return render(request, 'attendance/asset_tracking.html', {
        'assets': assets,
        'staff_users': staff_users,
    })


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse

@login_required
def stock_report(request):
    """Generate stock movement reports with Edit and Delete capabilities."""
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        action = request.POST.get('action')
        movement_id = request.POST.get('movement_id')
        movement = get_object_or_404(StockMovement, id=movement_id)

        if action == 'edit':
            movement.movement_type = request.POST.get('movement_type')
            movement.quantity = request.POST.get('quantity')
            movement.reference = request.POST.get('reference', '')
            movement.remarks = request.POST.get('remarks', '')
            movement.save()
            messages.success(request, "Stock movement updated successfully.")

        elif action == 'delete':
            movement.delete()
            messages.success(request, "Stock movement record deleted successfully.")

        return redirect(request.META.get('HTTP_REFERER', 'attendance:stock_report'))

    item_id = request.GET.get('item')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    movements = StockMovement.objects.select_related('item').all().order_by('-date')
    if item_id:
        movements = movements.filter(item_id=item_id)
    if start_date and end_date:
        movements = movements.filter(date__date__range=[start_date, end_date])

    items = InventoryItem.objects.all()
    return render(request, 'attendance/stock_report.html', {
        'movements': movements,
        'items': items,
        'selected_item': item_id,
        'start_date': start_date,
        'end_date': end_date,
    })


# ==================== TRANSPORT MODULE ====================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from .models import Vehicle, Route, TransportAllocation, TripLog

@login_required
def manage_vehicles(request):
    """CRUD for vehicles and routes with live frontend search & filtering."""
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        action = request.POST.get('action')

        # --- VEHICLE ACTIONS ---
        if action == 'add_vehicle':
            Vehicle.objects.create(
                registration_number=request.POST.get('registration_number'),
                model=request.POST.get('model'),
                capacity=request.POST.get('capacity'),
                driver_name=request.POST.get('driver_name'),
                driver_contact=request.POST.get('driver_contact'),
                status=request.POST.get('status')
            )
            messages.success(request, "Vehicle added successfully.")

        elif action == 'edit_vehicle':
            vehicle_id = request.POST.get('vehicle_id')
            vehicle = get_object_or_404(Vehicle, id=vehicle_id)
            vehicle.registration_number = request.POST.get('registration_number')
            vehicle.model = request.POST.get('model')
            vehicle.capacity = request.POST.get('capacity')
            vehicle.driver_name = request.POST.get('driver_name')
            vehicle.driver_contact = request.POST.get('driver_contact')
            vehicle.status = request.POST.get('status')
            vehicle.save()
            messages.success(request, f"Vehicle '{vehicle.registration_number}' updated successfully.")

        elif action == 'delete_vehicle':
            vehicle_id = request.POST.get('vehicle_id')
            vehicle = get_object_or_404(Vehicle, id=vehicle_id)
            reg_num = vehicle.registration_number
            vehicle.delete()
            messages.success(request, f"Vehicle '{reg_num}' deleted successfully.")

        # --- ROUTE ACTIONS ---
        elif action == 'add_route':
            vehicle_id = request.POST.get('vehicle')
            vehicle = Vehicle.objects.filter(id=vehicle_id).first() if vehicle_id else None
            Route.objects.create(
                name=request.POST.get('name'),
                vehicle=vehicle,
                start_location=request.POST.get('start_location'),
                end_location=request.POST.get('end_location'),
                departure_time=request.POST.get('departure_time'),
                arrival_time=request.POST.get('arrival_time'),
                days_of_week=request.POST.get('days_of_week')
            )
            messages.success(request, "Route added successfully.")

        elif action == 'edit_route':
            route_id = request.POST.get('route_id')
            route = get_object_or_404(Route, id=route_id)
            vehicle_id = request.POST.get('vehicle')
            route.vehicle = Vehicle.objects.filter(id=vehicle_id).first() if vehicle_id else None
            route.name = request.POST.get('name')
            route.start_location = request.POST.get('start_location')
            route.end_location = request.POST.get('end_location')
            route.departure_time = request.POST.get('departure_time')
            route.arrival_time = request.POST.get('arrival_time')
            route.days_of_week = request.POST.get('days_of_week')
            route.save()
            messages.success(request, f"Route '{route.name}' updated successfully.")

        elif action == 'delete_route':
            route_id = request.POST.get('route_id')
            route = get_object_or_404(Route, id=route_id)
            r_name = route.name
            route.delete()
            messages.success(request, f"Route '{r_name}' deleted successfully.")

        return redirect('attendance:manage_vehicles')

    vehicles = Vehicle.objects.all()
    routes = Route.objects.select_related('vehicle').all()

    return render(request, 'attendance/manage_vehicles.html', {
        'vehicles': vehicles,
        'routes': routes,
    })

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.contrib import messages
from .models import AcademicTerm, Route, StudentProfile, TransportAllocation, User
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.contrib import messages
from .models import AcademicTerm, Route, StudentProfile, TransportAllocation, User

@login_required
@transaction.atomic
def allocate_transport(request):
    """Assign students to routes per term with filtering support."""
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    terms = AcademicTerm.objects.all()
    routes = Route.objects.all()
    students = StudentProfile.objects.all()

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        route_id = request.POST.get('route_id')
        term_id = request.POST.get('term_id')
        
        student = get_object_or_404(StudentProfile, reg_number=student_id)
        route = get_object_or_404(Route, id=route_id)
        term = get_object_or_404(AcademicTerm, id=term_id)
        
        allocation, created = TransportAllocation.objects.get_or_create(
            student=student, term=term,
            defaults={'route': route, 'is_active': True}
        )
        if not created:
            allocation.route = route
            allocation.save()
            messages.info(request, "Allocation updated.")
        else:
            messages.success(request, "Student allocated to route.")
        return redirect('attendance:allocate_transport')

    allocations = TransportAllocation.objects.select_related('student', 'route', 'term').all()

    route_filter = request.GET.get('route', '')
    term_filter = request.GET.get('term', '')
    status_filter = request.GET.get('status', '')
    student_filter = request.GET.get('student', '')

    if route_filter:
        allocations = allocations.filter(route_id=route_filter)
    if term_filter:
        allocations = allocations.filter(term_id=term_filter)
    if status_filter:
        if status_filter == 'active':
            allocations = allocations.filter(is_active=True)
        elif status_filter == 'inactive':
            allocations = allocations.filter(is_active=False)
    if student_filter:
        allocations = allocations.filter(student__reg_number=student_filter)

    return render(request, 'attendance/allocate_transport.html', {
        'allocations': allocations,
        'terms': terms,
        'routes': routes,
        'students': students,
        'selected_route': route_filter,
        'selected_term': term_filter,
        'selected_status': status_filter,
        'selected_student': student_filter,
    })


@login_required
@transaction.atomic
def edit_transport_allocation(request, pk):
    """Edit an existing transport allocation."""
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    allocation = get_object_or_404(TransportAllocation, pk=pk)

    if request.method == 'POST':
        route_id = request.POST.get('route_id')
        term_id = request.POST.get('term_id')
        is_active = request.POST.get('is_active') == 'true'

        allocation.route = get_object_or_404(Route, id=route_id)
        allocation.term = get_object_or_404(AcademicTerm, id=term_id)
        allocation.is_active = is_active
        allocation.save()

        messages.success(request, "Transport allocation updated successfully.")
    
    return redirect('attendance:allocate_transport')


@login_required
@transaction.atomic
def delete_transport_allocation(request, pk):
    """Delete a transport allocation."""
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    allocation = get_object_or_404(TransportAllocation, pk=pk)

    if request.method == 'POST':
        allocation.delete()
        messages.success(request, "Transport allocation removed successfully.")

    return redirect('attendance:allocate_transport')

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.contrib import messages
from .models import TripLog, Vehicle, Route, User

@login_required
@transaction.atomic
def trip_log(request):
    """Log trips, update/delete existing logs, and view history."""
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        trip_id = request.POST.get('trip_id')

        # --- DELETE ACTION ---
        if action == 'delete':
            trip = get_object_or_404(TripLog, id=trip_id)
            trip.delete()
            messages.success(request, "Trip log deleted successfully.")
            return redirect('attendance:trip_log')

        # --- CREATE OR EDIT ACTION ---
        elif action in ['create', 'edit']:
            vehicle_id = request.POST.get('vehicle')
            route_id = request.POST.get('route')
            dep_time = request.POST.get('departure_time')
            arr_time = request.POST.get('arrival_time') or None
            driver = request.POST.get('driver_name')
            mileage_start = request.POST.get('mileage_start') or None
            mileage_end = request.POST.get('mileage_end') or None
            remarks = request.POST.get('remarks')

            vehicle = get_object_or_404(Vehicle, id=vehicle_id)
            route = get_object_or_404(Route, id=route_id)

            if action == 'edit':
                trip = get_object_or_404(TripLog, id=trip_id)
                trip.vehicle = vehicle
                trip.route = route
                trip.departure_time = dep_time
                trip.arrival_time = arr_time
                trip.driver_name = driver
                trip.mileage_start = mileage_start
                trip.mileage_end = mileage_end
                trip.remarks = remarks
                trip.save()
                messages.success(request, "Trip log updated successfully.")
            else:
                TripLog.objects.create(
                    vehicle=vehicle,
                    route=route,
                    departure_time=dep_time,
                    arrival_time=arr_time,
                    driver_name=driver,
                    mileage_start=mileage_start,
                    mileage_end=mileage_end,
                    remarks=remarks
                )
                messages.success(request, "Trip logged successfully.")

            return redirect('attendance:trip_log')

    # GET Request: Data retrieval
    trips = TripLog.objects.select_related('vehicle', 'route').all().order_by('-departure_time')
    vehicles = Vehicle.objects.all()
    routes = Route.objects.all()

    # Get unique list of drivers for filter dropdown
    drivers = TripLog.objects.exclude(driver_name__isnull=True)\
                             .exclude(driver_name__exact='')\
                             .values_list('driver_name', flat=True)\
                             .distinct()

    return render(request, 'attendance/trip_log.html', {
        'trips': trips,
        'vehicles': vehicles,
        'routes': routes,
        'drivers': drivers,
    })

# ==================== HUMAN RESOURCE MODULE (Admin) ====================

from .models import LeaveRequest, PerformanceEvaluation, Qualification

@login_required
def approve_leave(request, leave_id):
    """Approve or reject leave requests (Admin)."""
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    leave = get_object_or_404(LeaveRequest, id=leave_id)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            leave.status = 'APPROVED'
            leave.approved_by = request.user
            messages.success(request, f"Leave approved for {leave.staff.username}.")
        elif action == 'reject':
            leave.status = 'REJECTED'
            messages.info(request, f"Leave rejected for {leave.staff.username}.")
        leave.save()
    return redirect('attendance:apply_leave')


@login_required
def manage_performance(request):
    """Add/view/edit/delete performance evaluations."""
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    evaluations = PerformanceEvaluation.objects.select_related('staff', 'evaluator', 'term').all()
    staff_users = User.objects.exclude(role=User.IS_STUDENT)
    terms = AcademicTerm.objects.all()

    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Handle Deletion
        if action == 'delete':
            eval_id = request.POST.get('eval_id')
            eval_obj = get_object_or_404(PerformanceEvaluation, id=eval_id)
            eval_obj.delete()
            messages.success(request, "Evaluation deleted successfully.")
            return redirect('attendance:manage_performance')

        # Handle Editing
        elif action == 'edit':
            eval_id = request.POST.get('eval_id')
            eval_obj = get_object_or_404(PerformanceEvaluation, id=eval_id)
            
            staff_id = request.POST.get('staff')
            score = request.POST.get('score')
            comments = request.POST.get('comments')
            term_id = request.POST.get('term')
            
            eval_obj.staff = User.objects.get(id=staff_id)
            eval_obj.score = score
            eval_obj.comments = comments
            eval_obj.term = AcademicTerm.objects.get(id=term_id) if term_id else None
            eval_obj.save()
            
            messages.success(request, "Evaluation updated successfully.")
            return redirect('attendance:manage_performance')

        # Handle Creation (Default POST)
        else:
            staff_id = request.POST.get('staff')
            score = request.POST.get('score')
            comments = request.POST.get('comments')
            term_id = request.POST.get('term')
            staff = User.objects.get(id=staff_id)
            term = AcademicTerm.objects.get(id=term_id) if term_id else None
            
            PerformanceEvaluation.objects.create(
                staff=staff, evaluator=request.user, score=score,
                comments=comments, term=term
            )
            messages.success(request, "Evaluation added.")
            return redirect('attendance:manage_performance')

    return render(request, 'attendance/manage_performance.html', {
        'evaluations': evaluations,
        'staff_users': staff_users,
        'terms': terms,
    })


from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Qualification, User

@login_required
def manage_qualifications(request):
    """Add, view, edit, and delete staff qualifications."""
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    qualifications = Qualification.objects.select_related('staff').all()
    staff_users = User.objects.exclude(role=User.IS_STUDENT)

    if request.method == 'POST':
        action = request.POST.get('action')

        # Handle Deletion
        if action == 'delete':
            qual_id = request.POST.get('qual_id')
            qual = get_object_or_404(Qualification, id=qual_id)
            qual.delete()
            messages.success(request, "Qualification deleted successfully.")
            return redirect('attendance:manage_qualifications')

        # Handle Editing
        elif action == 'edit':
            qual_id = request.POST.get('qual_id')
            qual = get_object_or_404(Qualification, id=qual_id)

            staff_id = request.POST.get('staff')
            qual.staff = User.objects.get(id=staff_id)
            qual.qualification_name = request.POST.get('qualification_name')
            qual.institution = request.POST.get('institution')
            qual.year_awarded = request.POST.get('year_awarded')

            if 'certificate_file' in request.FILES:
                qual.certificate_file = request.FILES['certificate_file']

            qual.save()
            messages.success(request, "Qualification updated successfully.")
            return redirect('attendance:manage_qualifications')

        # Handle Creation (Default POST)
        else:
            staff_id = request.POST.get('staff')
            qual_name = request.POST.get('qualification_name')
            institution = request.POST.get('institution')
            year = request.POST.get('year_awarded')
            certificate = request.FILES.get('certificate_file')
            staff = User.objects.get(id=staff_id)

            Qualification.objects.create(
                staff=staff, qualification_name=qual_name,
                institution=institution, year_awarded=year,
                certificate_file=certificate
            )
            messages.success(request, "Qualification added.")
            return redirect('attendance:manage_qualifications')

    return render(request, 'attendance/manage_qualifications.html', {
        'qualifications': qualifications,
        'staff_users': staff_users,
    })

# ==================== HR MODULE (Comprehensive Dashboard View) ====================
import json
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Avg, Count, Q
from django.core.serializers.json import DjangoJSONEncoder
from .models import LeaveRequest, PerformanceEvaluation, Qualification, AcademicTerm, User

@login_required
def hr_dashboard(request):
    """
    Comprehensive HR Analytics & Logistics Overview.
    Aggregates data across Staff Profiles, Leave Requests, Performance Evaluations, and Qualifications.
    """
    if request.user.role not in [User.IS_ADMIN, User.IS_ACCOUNTANT]:
        return HttpResponse("Unauthorized", status=403)

    # --- 1. STAFF METRICS ---
    staff_qs = User.objects.exclude(role=User.IS_STUDENT)
    total_staff = staff_qs.count()
    
    # --- 2. LEAVE REQUEST METRICS ---
    total_leaves = LeaveRequest.objects.count()
    pending_leaves_count = LeaveRequest.objects.filter(status='PENDING').count()
    approved_leaves_count = LeaveRequest.objects.filter(status='APPROVED').count()
    rejected_leaves_count = LeaveRequest.objects.filter(status='REJECTED').count()

    # --- 3. PERFORMANCE EVALUATION METRICS ---
    total_evaluations = PerformanceEvaluation.objects.count()
    avg_performance = PerformanceEvaluation.objects.aggregate(Avg('score'))['score__avg'] or 0.0

    # Top performing staff members (by average score)
    top_performers = PerformanceEvaluation.objects.values(
        'staff__id', 'staff__username', 'staff__first_name', 'staff__last_name'
    ).annotate(
        avg_score=Avg('score'),
        total_evals=Count('id')
    ).order_by('-avg_score')[:5]

    # --- 4. QUALIFICATIONS METRICS ---
    total_qualifications = Qualification.objects.count()
    qualified_staff_count = Qualification.objects.values('staff').distinct().count()
    qualification_coverage = round((qualified_staff_count / total_staff * 100), 1) if total_staff > 0 else 0.0

    # --- 5. ACTIONABLE & RECENT LOGS ---
    # Pending leave requests requiring admin attention
    pending_leave_requests = LeaveRequest.objects.filter(
        status='PENDING'
    ).select_related('staff').order_by('-id')[:5]

    # Recent performance evaluations
    recent_evaluations = PerformanceEvaluation.objects.select_related(
        'staff', 'evaluator', 'term'
    ).order_by('-id')[:5]

    # Recent qualifications added
    recent_qualifications = Qualification.objects.select_related('staff').order_by('-id')[:5]

    # --- 6. CHART DATA ENCODING ---
    # Chart 1: Leave Request Breakdown
    leave_status_labels = ['Approved', 'Pending', 'Rejected']
    leave_status_data = [approved_leaves_count, pending_leaves_count, rejected_leaves_count]

    # Chart 2: Top Staff Performance Scores
    top_performer_labels = [
        f"{p['staff__first_name']} {p['staff__last_name']}".strip() or p['staff__username']
        for p in top_performers
    ]
    top_performer_data = [round(p['avg_score'], 1) for p in top_performers]

    context = {
        # Core KPIs
        'total_staff': total_staff,
        'total_leaves': total_leaves,
        'pending_leaves_count': pending_leaves_count,
        'approved_leaves_count': approved_leaves_count,
        'rejected_leaves_count': rejected_leaves_count,
        'avg_performance': round(avg_performance, 1),
        'total_evaluations': total_evaluations,
        'total_qualifications': total_qualifications,
        'qualified_staff_count': qualified_staff_count,
        'qualification_coverage': qualification_coverage,

        # Recent Lists & Tables
        'pending_leave_requests': pending_leave_requests,
        'top_performers': top_performers,
        'recent_evaluations': recent_evaluations,
        'recent_qualifications': recent_qualifications,

        # JSON Payloads for Charts
        'leave_status_labels_json': json.dumps(leave_status_labels, cls=DjangoJSONEncoder),
        'leave_status_data_json': json.dumps(leave_status_data, cls=DjangoJSONEncoder),
        'performer_labels_json': json.dumps(top_performer_labels, cls=DjangoJSONEncoder),
        'performer_data_json': json.dumps(top_performer_data, cls=DjangoJSONEncoder),
    }

    return render(request, 'attendance/hr_dashboard.html', context)


# ==================== PARENT PORTAL (Admin/Registrar) ====================

from .models import ParentProfile

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import ParentProfile, StudentProfile, User

@login_required
def link_parent_student(request):
    """Admin/Registrar links parent accounts to students."""
    if request.user.role not in [User.IS_ADMIN, User.IS_REGISTRAR]:
        return HttpResponse("Unauthorized", status=403)

    # Use prefetch_related for optimal rendering of student lists in the template
    parents = ParentProfile.objects.select_related('user').prefetch_related('students').all()
    students = StudentProfile.objects.all()

    if request.method == 'POST':
        action = request.POST.get('action', 'link')
        
        if action == 'unlink_all':
            parent_id = request.POST.get('parent_id')
            parent = get_object_or_404(ParentProfile, id=parent_id)
            parent.students.clear()
            messages.success(request, f"Unlinked all students from {parent.name}.")
            return redirect('attendance:link_parent_student')

        else: # Default: link / update students
            parent_id = request.POST.get('parent_id')
            student_ids = request.POST.getlist('student_ids')
            parent = get_object_or_404(ParentProfile, id=parent_id)
            parent.students.set(student_ids)
            parent.save()
            messages.success(request, f"Updated student links for {parent.name}.")
            return redirect('attendance:link_parent_student')

    return render(request, 'attendance/link_parent_student.html', {
        'parents': parents,
        'students': students,
    })


# ==================== REGISTRAR-SPECIFIC VIEWS ====================

@login_required
def manage_admissions(request):
    """Process student admissions (create student accounts)."""
    if request.user.role not in [User.IS_ADMIN, User.IS_REGISTRAR]:
        return HttpResponse("Unauthorized", status=403)

    # Similar to manage_students but with admission-specific fields
    # Reuse existing manage_students logic or add enhancements
    # For brevity, we'll forward to the existing view (or implement here)
    return redirect('attendance:manage_students')


from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from .models import StudentProfile, Stream, Course, User

@login_required
@transaction.atomic
def manage_student_transfers(request):
    """Transfer students between streams/courses or delete student records."""
    if request.user.role not in [User.IS_ADMIN, User.IS_REGISTRAR]:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        action = request.POST.get('action', 'transfer')
        student_id = request.POST.get('student_id')
        student = get_object_or_404(StudentProfile, reg_number=student_id)

        if action == 'delete':
            student_name = student.name
            student.delete()
            messages.success(request, f"Student record for {student_name} deleted successfully.")
            return redirect('attendance:manage_student_transfers')

        else:  # Default transfer action
            new_stream_id = request.POST.get('new_stream')
            new_course_code = request.POST.get('new_course')
            
            if new_stream_id:
                student.stream = Stream.objects.get(id=new_stream_id)
            if new_course_code:
                student.course = Course.objects.get(code=new_course_code)
                
            student.save()
            messages.success(request, f"Student {student.name} transferred successfully.")
            return redirect('attendance:manage_student_transfers')

    students = StudentProfile.objects.select_related('stream', 'course').all()
    streams = Stream.objects.all()
    courses = Course.objects.all()
    
    return render(request, 'attendance/manage_student_transfers.html', {
        'students': students,
        'streams': streams,
        'courses': courses,
    })